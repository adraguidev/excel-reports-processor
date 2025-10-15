import os
import pandas as pd
import xlwt
from typing import Callable, Optional, List
import numpy as np
from concurrent.futures import ThreadPoolExecutor
import gc
import time
from concurrent.futures import ProcessPoolExecutor
import openpyxl
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import xlsxwriter
import multiprocessing as mp
from multiprocessing import Pool
import shutil


class ExcelProcessor:
    """Class to handle Excel file processing operations."""
    
    def __init__(self, base_dir: str = "ASIGNACIONES"):
        """Initialize the processor with base directory for files."""
        self.base_dir = base_dir
        self.file_types = ['CCM', 'PRR']

    def process_file(self, file_type: str, progress_callback: Optional[Callable[[str], None]] = None) -> bool:
        """
        Process a single Excel file of the specified type.
        
        Args:
            file_type: Type of file to process (CCM, PRR, or SOL)
            progress_callback: Optional callback function to report progress
        
        Returns:
            bool: True if processing was successful, False otherwise
        """
        try:
            if progress_callback:
                progress_callback(f"Processing {file_type} file...")

            input_path = os.path.join(self.base_dir, f"{file_type}.xls")
            output_path = os.path.join(self.base_dir, f"{file_type}-PROCESADO.xls")

            # Read the Excel file with xlrd engine
            df = pd.read_excel(input_path, engine='xlrd')

            if progress_callback:
                progress_callback(f"Filtering {file_type} data...")

            # Filter rows where column F (TRAMITE) starts with 'LM'
            df_filtered = df[df.iloc[:, 5].astype(str).str.startswith('LM', na=False)].copy()

            # Create new DataFrame with TRAMITE, OPERADOR, and FECHA_ASIGNACION columns
            df_filtered['TRAMITE'] = df_filtered.iloc[:, 5]  # Column F
            # Standard column mapping for CCM and PRR
            df_filtered['OPERADOR'] = df_filtered.iloc[:, 32]  # Column AG
            df_filtered['FECHA_ASIGNACION'] = pd.to_datetime(df_filtered.iloc[:, 39], dayfirst=True)  # Column AN
            
            # Format date as dd/mm/yyyy without time component
            df_filtered['FECHA_ASIGNACION'] = pd.to_datetime(df_filtered['FECHA_ASIGNACION'], dayfirst=True).dt.date
            df_processed = df_filtered[['TRAMITE', 'OPERADOR', 'FECHA_ASIGNACION']]

            if progress_callback:
                progress_callback(f"Filtered {len(df_processed)} records with LM TRAMITE. Saving processed {file_type} file...")

            # Save the processed DataFrame directly to Excel XLSX format with date formatting
            output_path = os.path.join(self.base_dir, f"{file_type}-PROCESADO.xlsx")
            with pd.ExcelWriter(output_path, engine='openpyxl', date_format='dd/mm/yyyy', datetime_format='dd/mm/yyyy') as writer:
                # Ensure date format is consistent
                df_processed['FECHA_ASIGNACION'] = pd.to_datetime(df_processed['FECHA_ASIGNACION']).dt.date
                df_processed.to_excel(writer, index=False)
                
                # Get the worksheet and set the number format for the date column
                worksheet = writer.sheets['Sheet1']
                for idx, col in enumerate(df_processed.columns):
                    if col == 'FECHA_ASIGNACION':
                        for cell in worksheet[f"{chr(65 + idx)}2:{chr(65 + idx)}{len(df_processed) + 1}"]:
                            cell[0].number_format = 'dd/mm/yyyy'

            if progress_callback:
                progress_callback(f"Successfully processed {file_type} file")

            return True

        except Exception as e:
            if progress_callback:
                progress_callback(f"Error processing {file_type} file: {str(e)}")
            return False

    def process_calidades_file(self, file_type: str, progress_callback: Optional[Callable[[str], None]] = None) -> bool:
        """
        Process a single CALIDADES Excel file of the specified type.
        
        Args:
            file_type: Type of file to process (CCM, PRR, or SOL)
            progress_callback: Optional callback function to report progress
        
        Returns:
            bool: True if processing was successful, False otherwise
        """
        try:
            if progress_callback:
                progress_callback(f"Processing {file_type}-CALIDADES file...")

            input_path = os.path.join(self.base_dir, f"{file_type}-CALIDADES.xls")
            output_path = os.path.join(self.base_dir, f"{file_type}-CALIDADES-PROCESADO.xls")

            # Read the Excel file
            df = pd.read_excel(input_path, engine='xlrd')

            if progress_callback:
                progress_callback(f"Analyzing rows in {file_type}-CALIDADES...")

            # Lists to store processed data
            tramites = []
            valores = []

            # Determine TRAMITE column (Column R for CCM and PRR)
            tramite_col = 17  # Column R for CCM and PRR
            
            # Iterate through DataFrame rows
            idx = 0
            while idx < len(df):
                # Get value from appropriate TRAMITE column
                tramite_value = str(df.iloc[idx, tramite_col]).strip()
                
                # Check if the value starts with 'LM'
                if tramite_value.startswith('LM'):
                    # Get the value from column AK (index 36)
                    ak_value = str(df.iloc[idx, 36]).strip()
                    
                    if ak_value and ak_value not in ['-', '- ']:
                        # Use the AK value directly
                        tramites.append(tramite_value)
                        valores.append(ak_value)
                    else:
                        # Check if we can access the row two rows below
                        if idx + 2 < len(df):
                            # Get value from column G (index 6) two rows below
                            g_value = str(df.iloc[idx + 2, 6]).strip()
                            tramites.append(tramite_value)
                            valores.append(g_value)
                
                idx += 1

            if progress_callback:
                progress_callback(f"Found and processed {len(tramites)} LM records from column R (TRAMITE). Saving processed {file_type}-CALIDADES file...")

            # Create a new workbook and add a worksheet
            wb = xlwt.Workbook()
            ws = wb.add_sheet('Sheet1')

            # Write headers
            headers = ['TRAMITE', 'VALOR']
            for col, header in enumerate(headers):
                ws.write(0, col, header)

            # Write data rows
            for row_idx, (tramite, valor) in enumerate(zip(tramites, valores)):
                ws.write(row_idx + 1, 0, tramite)
                ws.write(row_idx + 1, 1, valor if valor else '')  # Ensure empty values are written as empty strings

            # Save the workbook
            wb.save(output_path)

            if progress_callback:
                progress_callback(f"Successfully processed {file_type}-CALIDADES file with {len(tramites)} records")

            return True

        except Exception as e:
            if progress_callback:
                progress_callback(f"Error processing {file_type}-CALIDADES file: {str(e)}")
            return False

    def process_all_calidades_files(self, progress_callback: Optional[Callable[[str], None]] = None) -> List[str]:
        """
        Process all CALIDADES Excel files (CCM, PRR).
        
        Args:
            progress_callback: Optional callback function to report progress
        
        Returns:
            List[str]: List of successfully processed file types
        """
        successful_files = []
        
        for file_type in self.file_types:
            if self.process_calidades_file(file_type, progress_callback):
                successful_files.append(file_type)
        
        if progress_callback:
            if successful_files:
                progress_callback(f"Completed processing CALIDADES files. Successfully processed: {', '.join(successful_files)}")
            else:
                progress_callback("Completed processing CALIDADES files. No files were processed successfully.")
        
        return successful_files
        
    def process_cross_files(self, progress_callback: Optional[Callable[[str], None]] = None) -> List[str]:
        """
        Process cross-referencing between consolidated CSV files and processed XLSX files.
        
        Args:
            progress_callback: Optional callback function to report progress
         
        Returns:
            List[str]: List of successfully processed file types
        """
        # First process the regular cross-referencing
        successful_files = self._process_regular_cross_files(progress_callback)
        
        # Then process the personnel cross-referencing
        if successful_files:
            self._process_personnel_cross_files(progress_callback)
        
        return successful_files
        
    def _process_personnel_cross_files(self, progress_callback: Optional[Callable[[str], None]] = None) -> None:
        """Process cross-referencing with PERSONAL.xlsx file"""
        try:
            if progress_callback:
                progress_callback("Starting personnel cross-referencing...")
            
            # Read PERSONAL.xlsx
            personal_path = os.path.join(self.base_dir, 'PERSONAL.xlsx')
            if not os.path.exists(personal_path):
                if progress_callback:
                    progress_callback(f"Warning: {personal_path} not found. Skipping personnel cross-reference.")
                return
            
            # Read PERSONAL.xlsx and keep all columns
            df_personal = pd.read_excel(personal_path)
            if progress_callback:
                progress_callback(f"Loaded PERSONAL.xlsx with {len(df_personal.columns)} columns")
            
            # Process each consolidated file
            for file_type in ['CCM', 'PRR']:
                try:
                    # Read consolidated final file
                    consolidated_path = os.path.join('descargas', file_type, f'consolidado_final_{file_type}.xlsx')
                    if not os.path.exists(consolidated_path):
                        if progress_callback:
                            progress_callback(f"Warning: {consolidated_path} not found. Skipping {file_type}.")
                        continue
                    
                    if progress_callback:
                        progress_callback(f"Processing personnel cross-reference for {file_type}...")
                    
                    # Read consolidated file
                    df_consolidated = pd.read_excel(consolidated_path)
                    initial_columns = df_consolidated.columns.tolist()
                    
                    if progress_callback:
                        progress_callback(f"Merging {file_type} data with PERSONAL information...")
                    
                    # Convert OPERADOR column to uppercase for case-insensitive matching
                    df_consolidated['OPERADOR'] = df_consolidated['OPERADOR'].str.upper()

                    # Merge with personnel data using 'APELLIDOS Y NOMBRES' column
                    # Keep all columns from both DataFrames
                    df_merged = pd.merge(
                        df_consolidated,
                        df_personal,
                        left_on='OPERADOR',
                        right_on='APELLIDOS Y NOMBRES',
                        how='left'
                    )
                    
                    # Remove duplicate APELLIDOS Y NOMBRES column if it exists
                    if 'APELLIDOS Y NOMBRES' in df_merged.columns and 'OPERADOR' in df_merged.columns:
                        df_merged = df_merged.drop('APELLIDOS Y NOMBRES', axis=1)
                    
                    # Save merged result
                    output_path = os.path.join('descargas', file_type, f'consolidado_final_{file_type}_personal.xlsx')
                    with pd.ExcelWriter(output_path, engine='openpyxl', date_format='dd/mm/yyyy', datetime_format='dd/mm/yyyy') as writer:
                        df_merged.to_excel(writer, index=False)
                        
                        # Format date columns
                        worksheet = writer.sheets['Sheet1']
                        for idx, col in enumerate(df_merged.columns):
                            if 'FECHA' in col.upper():
                                for cell in worksheet[f"{chr(65 + idx)}2:{chr(65 + idx)}{len(df_merged) + 1}"]:
                                    cell[0].number_format = 'dd/mm/yyyy'
                    
                    # Log the number of columns added from PERSONAL.xlsx
                    new_columns = [col for col in df_merged.columns if col not in initial_columns]
                    if progress_callback:
                        progress_callback(f"Added {len(new_columns)} columns from PERSONAL.xlsx to {file_type}")
                        progress_callback(f"Successfully created personnel cross-reference for {file_type}")
                    
                except Exception as e:
                    if progress_callback:
                        progress_callback(f"Error processing personnel cross-reference for {file_type}: {str(e)}")
                    continue
            
            if progress_callback:
                progress_callback("Completed personnel cross-reference processing.")
                
        except Exception as e:
            if progress_callback:
                progress_callback(f"Error in personnel cross-reference processing: {str(e)}")


    def _process_regular_cross_files(self, progress_callback: Optional[Callable[[str], None]] = None) -> List[str]:
        """Internal method to handle the regular cross-file processing"""

        successful_files = []
        
        for file_type in self.file_types:
            try:
                if progress_callback:
                    progress_callback(f"Processing cross-reference for {file_type}...")
                
                # Read consolidated CSV file
                csv_path = os.path.join('descargas', file_type, f'consolidado_total_{file_type}.csv')
                if not os.path.exists(csv_path):
                    if progress_callback:
                        progress_callback(f"Warning: {csv_path} not found. Skipping {file_type}")
                    continue
                
                df_csv = pd.read_csv(csv_path, low_memory=False, dtype={'8': str, '9': str, '10': str})
                csv_record_count = len(df_csv)
                if progress_callback:
                    progress_callback(f"Initial CSV record count for {file_type}: {csv_record_count}")
                
                # Read processed XLSX file
                xlsx_path = os.path.join(self.base_dir, f'{file_type}-PROCESADO.xlsx')
                if not os.path.exists(xlsx_path):
                    if progress_callback:
                        progress_callback(f"Warning: {xlsx_path} not found. Skipping {file_type}")
                    continue
                
                df_xlsx = pd.read_excel(xlsx_path)
                # Remove duplicates from XLSX before merging, keeping first occurrence
                df_xlsx = df_xlsx.drop_duplicates(subset=['TRAMITE'], keep='first')
                xlsx_record_count = len(df_xlsx)
                if progress_callback:
                    progress_callback(f"XLSX record count for {file_type} (after removing duplicates): {xlsx_record_count}")
                
                if progress_callback:
                    progress_callback(f"Merging {file_type} files...")
                
                # Merge DataFrames on NumeroTramite and TRAMITE columns
                df_merged = pd.merge(
                    df_csv,
                    df_xlsx[['TRAMITE', 'OPERADOR', 'FECHA_ASIGNACION']],
                    left_on='NumeroTramite',
                    right_on='TRAMITE',
                    how='left'
                )
                
                # Check for record count discrepancies
                merged_record_count = len(df_merged)
                if merged_record_count != csv_record_count:
                    if progress_callback:
                        progress_callback(f"Warning: Record count mismatch in {file_type}:")
                        progress_callback(f"  - Initial CSV records: {csv_record_count}")
                        progress_callback(f"  - Final merged records: {merged_record_count}")
                        
                        # Analyze discrepancies
                        if merged_record_count < csv_record_count:
                            missing_records = csv_record_count - merged_record_count
                            progress_callback(f"  - {missing_records} records were lost during merge")
                            # Check for duplicate NumeroTramite in CSV
                            csv_duplicates = df_csv['NumeroTramite'].duplicated().sum()
                            if csv_duplicates > 0:
                                progress_callback(f"  - Found {csv_duplicates} duplicate NumeroTramite entries in CSV")
                        else:
                            extra_records = merged_record_count - csv_record_count
                            progress_callback(f"  - {extra_records} additional records were created during merge")
                            # Check for duplicate matches
                            merge_duplicates = df_merged['NumeroTramite'].duplicated().sum()
                            if merge_duplicates > 0:
                                progress_callback(f"  - Found {merge_duplicates} duplicate entries after merge")
                
                # Remove duplicate TRAMITE column
                df_merged = df_merged.drop('TRAMITE', axis=1)
                
                # Save merged result as XLSX
                xlsx_output_path = os.path.join('descargas', file_type, f'consolidado_final_{file_type}.xlsx')
                with pd.ExcelWriter(xlsx_output_path, engine='openpyxl', date_format='dd/mm/yyyy', datetime_format='dd/mm/yyyy') as writer:
                    df_merged.to_excel(writer, index=False)
                    
                    # Get the worksheet and set the number format for date columns
                    worksheet = writer.sheets['Sheet1']
                    for idx, col in enumerate(df_merged.columns):
                        if 'FECHA' in col.upper():
                            for cell in worksheet[f"{chr(65 + idx)}2:{chr(65 + idx)}{len(df_merged) + 1}"]:
                                cell[0].number_format = 'dd/mm/yyyy'
                
                # Save merged result as CSV
                csv_output_path = os.path.join('descargas', file_type, f'consolidado_final_{file_type}.csv')
                df_merged.to_csv(csv_output_path, index=False)
                
                if progress_callback:
                    progress_callback(f"Successfully processed cross-reference for {file_type}")
                
                successful_files.append(file_type)
                
            except Exception as e:
                if progress_callback:
                    progress_callback(f"Error processing cross-reference for {file_type}: {str(e)}")
                continue
        
        if progress_callback:
            if successful_files:
                progress_callback(f"Completed cross-reference processing. Successfully processed: {', '.join(successful_files)}")
            else:
                progress_callback("Completed cross-reference processing. No files were processed successfully.")
        
        return successful_files

    def process_all_files(self, progress_callback: Optional[Callable[[str], None]] = None) -> List[str]:
        """
        Process all Excel files (CCM, PRR).
        
        Args:
            progress_callback: Optional callback function to report progress
        
        Returns:
            List[str]: List of successfully processed file types
        """
        successful_files = []
        
        for file_type in self.file_types:
            if self.process_file(file_type, progress_callback):
                successful_files.append(file_type)
        
        if progress_callback:
            if successful_files:
                progress_callback(f"Completed processing. Successfully processed: {', '.join(successful_files)}")
            else:
                progress_callback("Completed processing. No files were processed successfully.")
        
        return successful_files
        
    def optimized_format_files(self, progress_callback: Optional[Callable[[str], None]] = None) -> List[str]:
        """
        Método optimizado para formatear archivos Excel.
        Mantiene funcionalidad EXACTA para FechaPre pero optimizado para velocidad.
        
        Returns:
            List[str]: Lista de archivos procesados exitosamente
        """
        import pandas as pd  # Asegurar que pandas está disponible
        successful_files = []
        
        def format_single_file(file_type):
            """Formatear un archivo con lógica correcta pero optimizada"""
            try:
                # Leer archivo personal existente
                file_path = os.path.join('descargas', file_type, f'consolidado_final_{file_type}_personal.xlsx')
                if not os.path.exists(file_path):
                    if progress_callback:
                        progress_callback(f"⚠️ Archivo no encontrado: {file_path}")
                    return None
                
                if progress_callback:
                    progress_callback(f"Formateando archivo {file_type}...")
                
                # 1. Leer archivo Excel
                df = pd.read_excel(file_path)
                
                if progress_callback:
                    progress_callback(f"📊 {file_type}: {len(df.columns)} columnas, {len(df)} filas")
                
                # 2. Definir columnas de fecha 
                date_columns = ['FechaExpendiente', 'FechaEtapaAprobacionMasivaFin', 'FechaPre', 'FECHA_ASIGNACION']
                
                # 3. OPTIMIZACIÓN: Pre-procesar columnas de fecha vectorialmente pero manteniendo la lógica original
                if progress_callback:
                    progress_callback(f"⚡ Pre-procesando fechas con lógica original...")
                
                df_processed = df.copy()
                fechapre_converted = 0
                
                for col_name in date_columns:
                    if col_name in df_processed.columns:
                        def convert_date_value(value):
                            nonlocal fechapre_converted
                            try:
                                # LÓGICA EXACTA del método original
                                if pd.notna(value):
                                    if isinstance(value, (datetime, pd.Timestamp)):
                                        date_value = value
                                    else:
                                        # Try parsing string dates (igual que el original)
                                        try:
                                            date_value = pd.to_datetime(str(value), dayfirst=True)
                                        except:
                                            try:
                                                date_value = pd.to_datetime(str(value))
                                            except:
                                                return value  # Si no puede convertir, devolver original
                                    
                                    # Convert to Excel serial number (igual que el original)
                                    excel_date = date_value.toordinal() - datetime(1900, 1, 1).toordinal() + 2
                                    
                                    # Contar conversiones de FechaPre
                                    if col_name == 'FechaPre':
                                        fechapre_converted += 1
                                    
                                    return excel_date
                                else:
                                    return value
                            except:
                                return value  # En caso de cualquier error, devolver original
                        
                        # Aplicar conversión vectorizada
                        df_processed[col_name] = df_processed[col_name].apply(convert_date_value)
                
                if progress_callback:
                    progress_callback(f"🔍 {fechapre_converted} celdas de FechaPre convertidas correctamente")
                
                # 4. OPTIMIZACIÓN: Escritura rápida con openpyxl
                if progress_callback:
                    progress_callback(f"💾 Escritura optimizada...")
                
                wb = openpyxl.Workbook()
                ws = wb.active
                
                # Escribir headers rápidamente
                for col_idx, col_name in enumerate(df_processed.columns, 1):
                    ws.cell(row=1, column=col_idx, value=col_name)
                
                # OPTIMIZACIÓN: Escribir datos por bloques
                data_rows = df_processed.values.tolist()
                for row_idx, row_data in enumerate(data_rows, 2):
                    for col_idx, value in enumerate(row_data):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 5. OPTIMIZACIÓN: Aplicar formatos por columnas completas
                if progress_callback:
                    progress_callback(f"🎨 Aplicando formatos...")
                
                for col_idx, col_name in enumerate(df_processed.columns, 1):
                    col_letter = get_column_letter(col_idx)
                    
                    if col_name in date_columns:
                        # Aplicar formato de fecha a toda la columna
                        range_str = f"{col_letter}2:{col_letter}{len(df_processed) + 1}"
                        for cell in ws[range_str]:
                            if isinstance(cell, tuple):
                                for c in cell:
                                    c.number_format = 'dd/mm/yyyy'
                            else:
                                cell.number_format = 'dd/mm/yyyy'
                    elif df_processed[col_name].dtype in ['int64', 'float64', 'int32', 'float32']:
                        # Aplicar formato numérico
                        range_str = f"{col_letter}2:{col_letter}{len(df_processed) + 1}"
                        for cell in ws[range_str]:
                            if isinstance(cell, tuple):
                                for c in cell:
                                    c.number_format = '#,##0'
                            else:
                                cell.number_format = '#,##0'
                
                # 6. OPTIMIZACIÓN: Anchos calculados rápidamente
                if progress_callback:
                    progress_callback(f"📏 Calculando anchos...")
                
                for col_idx, col_name in enumerate(df_processed.columns, 1):
                    col_letter = get_column_letter(col_idx)
                    
                    # Cálculo rápido de ancho
                    if df_processed[col_name].dtype == 'object':
                        max_len = df_processed[col_name].astype(str).str.len().max()
                        header_len = len(str(col_name))
                        optimal_width = min(max(max_len, header_len) + 2, 50)
                    else:
                        optimal_width = min(max(15, len(str(col_name)) + 2), 50)
                    
                    ws.column_dimensions[col_letter].width = optimal_width
                
                # 7. Crear tabla
                if progress_callback:
                    progress_callback(f"📋 Creando tabla...")
                
                tab = Table(displayName="BASE", ref=f"A1:{get_column_letter(len(df_processed.columns))}{len(df_processed) + 1}")
                style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                tab.tableStyleInfo = style
                ws.add_table(tab)
                
                # 8. Guardar archivo
                if progress_callback:
                    progress_callback(f"💾 Guardando...")
                
                wb.save(file_path)
                
                # Liberación de memoria
                del df, df_processed, data_rows, wb, ws
                gc.collect()
                
                if progress_callback:
                    progress_callback(f"✅ {file_type} formateado exitosamente y rápido.")
                
                return file_type
                
            except Exception as e:
                if progress_callback:
                    progress_callback(f"❌ Error al procesar {file_type}: {str(e)}")
                return None
        
        try:
            if progress_callback:
                progress_callback(f"🚀 Iniciando formateo optimizado (correcto Y rápido)...")
            
            # OPTIMIZACIÓN: Procesamiento en paralelo
            with ThreadPoolExecutor(max_workers=2) as executor:
                futures = {executor.submit(format_single_file, file_type): file_type 
                          for file_type in ['CCM', 'PRR']}
                
                for future in futures:
                    result = future.result()
                    if result:
                        successful_files.append(result)
            
            if progress_callback:
                if successful_files:
                    progress_callback(f"🎉 Formateo optimizado completado!")
                    progress_callback(f"📁 Archivos: {', '.join(successful_files)}")
                    progress_callback(f"⚡ FechaPre convertida correctamente Y con alta velocidad")
                else:
                    progress_callback("⚠️ No se pudo formatear ningún archivo")
            
        except Exception as e:
            if progress_callback:
                progress_callback(f"❌ Error en formateo: {str(e)}")
        
        return successful_files

    def ultra_fast_cross_processing(self, progress_callback: Optional[Callable[[str], None]] = None) -> List[str]:
        """
        Método ULTRA-OPTIMIZADO para procesamiento de cruces.
        Aplica todas las optimizaciones extremas: vectorización + xlsxwriter + threading.
        Mantiene TODA la funcionalidad del método original pero con máxima velocidad.
        
        Returns:
            List[str]: Lista de archivos procesados exitosamente
        """
        import pandas as pd
        import xlsxwriter
        from datetime import datetime
        import time
        
        start_time = time.time()
        successful_files = []
        
        if progress_callback:
            progress_callback("🚀 Iniciando procesamiento de cruces ULTRA-OPTIMIZADO...")
            progress_callback("⚡ Usando: vectorización + xlsxwriter + threading + limpieza avanzada")
        
        def process_single_cross_ultra(file_type):
            """Procesar cruce de un archivo con optimizaciones extremas"""
            try:
                if progress_callback:
                    progress_callback(f"🔄 Procesando cruce ULTRA para {file_type}...")
                
                # 1. LECTURA ULTRA-OPTIMIZADA con tipos específicos
                csv_path = os.path.join('descargas', file_type, f'consolidado_total_{file_type}.csv')
                xlsx_path = os.path.join(self.base_dir, f'{file_type}-PROCESADO.xlsx')
                
                if not os.path.exists(csv_path) or not os.path.exists(xlsx_path):
                    if progress_callback:
                        progress_callback(f"⚠️ Archivos faltantes para {file_type}")
                    return None
                
                # Lectura optimizada con chunks y tipos
                df_csv = pd.read_csv(csv_path, 
                                   dtype={'8': str, '9': str, '10': str}, 
                                   low_memory=False)
                df_xlsx = pd.read_excel(xlsx_path)
                
                if progress_callback:
                    progress_callback(f"📊 {file_type}: CSV {len(df_csv)} filas, XLSX {len(df_xlsx)} filas")
                
                # 2. OPTIMIZACIÓN EXTREMA: Limpiar duplicados vectorialmente
                df_xlsx_clean = df_xlsx.drop_duplicates(subset=['TRAMITE'], keep='first')
                
                # 3. MERGE ULTRA-OPTIMIZADO con índices
                if progress_callback:
                    progress_callback(f"🔗 Realizando merge vectorizado para {file_type}...")
                
                # Crear índice temporal para merge más rápido
                df_xlsx_clean.set_index('TRAMITE', inplace=True)
                
                # Merge vectorizado
                df_merged = pd.merge(
                    df_csv,
                    df_xlsx_clean[['OPERADOR', 'FECHA_ASIGNACION']],
                    left_on='NumeroTramite',
                    right_index=True,
                    how='left'
                )
                
                # 4. ESCRITURA ULTRA-RÁPIDA con xlsxwriter
                if progress_callback:
                    progress_callback(f"💾 Escritura ultra-rápida para {file_type}...")
                
                # Archivo base (sin personal)
                output_path_base = os.path.join('descargas', file_type, f'consolidado_final_{file_type}.xlsx')
                temp_path_base = output_path_base.replace('.xlsx', '_temp_base.xlsx')
                
                # Configurar xlsxwriter para velocidad máxima
                workbook_base = xlsxwriter.Workbook(temp_path_base, {
                    'constant_memory': True,
                    'nan_inf_to_errors': True,
                    'tmpdir': '.',
                })
                worksheet_base = workbook_base.add_worksheet('BASE')
                
                # Formatos pre-definidos
                date_format = workbook_base.add_format({'num_format': 'dd/mm/yyyy'})
                
                # 5. LIMPIEZA VECTORIZADA de datos
                df_merged = df_merged.replace([np.inf, -np.inf], np.nan)
                df_merged = df_merged.fillna('')
                
                # 6. ESCRITURA MASIVA por chunks
                # Headers
                for col_idx, col_name in enumerate(df_merged.columns):
                    worksheet_base.write(0, col_idx, col_name)
                
                # Identificar columnas de fecha vectorialmente
                date_cols = [col for col in df_merged.columns if 'FECHA' in col.upper()]
                date_col_indices = [df_merged.columns.get_loc(col) for col in date_cols]
                
                # Escritura por chunks ultra-rápida
                chunk_size = 8000
                total_rows = len(df_merged)
                
                for start_idx in range(0, total_rows, chunk_size):
                    end_idx = min(start_idx + chunk_size, total_rows)
                    chunk_data = df_merged.iloc[start_idx:end_idx].values
                    
                    for row_offset, row_data in enumerate(chunk_data):
                        row_idx = start_idx + row_offset + 1
                        
                        for col_idx, value in enumerate(row_data):
                            if col_idx in date_col_indices and pd.notna(value) and value != '':
                                worksheet_base.write(row_idx, col_idx, value, date_format)
                            else:
                                safe_value = str(value) if pd.notna(value) and value != '' else ''
                                worksheet_base.write(row_idx, col_idx, safe_value)
                
                # 7. Configurar tabla y anchos automáticamente
                last_col = xlsxwriter.utility.xl_col_to_name(len(df_merged.columns) - 1)
                table_range = f'A1:{last_col}{len(df_merged) + 1}'
                
                worksheet_base.add_table(table_range, {
                    'name': 'BASE',
                    'style': 'Table Style Medium 2',
                    'first_column': False,
                    'last_column': False,
                    'banded_rows': True,
                    'banded_columns': False,
                })
                
                # Anchos automáticos optimizados
                for col_idx in range(len(df_merged.columns)):
                    col_name = df_merged.columns[col_idx]
                    if df_merged[col_name].dtype == 'object':
                        sample_data = df_merged[col_name].dropna().astype(str).head(500)
                        if len(sample_data) > 0:
                            max_len = sample_data.str.len().max()
                            optimal_width = min(max(max_len, len(col_name)) + 2, 40)
                        else:
                            optimal_width = len(col_name) + 2
                    else:
                        optimal_width = max(12, len(col_name) + 2)
                    
                    worksheet_base.set_column(col_idx, col_idx, optimal_width)
                
                workbook_base.close()
                
                # Mover archivo base
                shutil.move(temp_path_base, output_path_base)
                
                # También guardar CSV para compatibilidad
                csv_output_path = os.path.join('descargas', file_type, f'consolidado_final_{file_type}.csv')
                df_merged.to_csv(csv_output_path, index=False)
                
                if progress_callback:
                    progress_callback(f"✅ Cruce base completado para {file_type}")
                
                return file_type
                
            except Exception as e:
                # Limpiar archivos temporales
                temp_path_base = os.path.join('descargas', file_type, f'consolidado_final_{file_type}_temp_base.xlsx')
                if os.path.exists(temp_path_base):
                    os.remove(temp_path_base)
                
                if progress_callback:
                    progress_callback(f"❌ Error en cruce ultra para {file_type}: {str(e)}")
                return None
        
        try:
            # 1. PASO 1: Pre-cargar archivo PERSONAL una sola vez y optimizarlo
            personal_path = os.path.join(self.base_dir, 'PERSONAL.xlsx')
            df_personal = None
            
            if os.path.exists(personal_path):
                if progress_callback:
                    progress_callback("📊 Cargando y optimizando archivo PERSONAL...")
                
                df_personal = pd.read_excel(personal_path, engine='openpyxl')
                # Optimización vectorizada de strings
                df_personal['APELLIDOS Y NOMBRES'] = df_personal['APELLIDOS Y NOMBRES'].str.upper().str.strip()
                
                if progress_callback:
                    progress_callback(f"📋 PERSONAL cargado: {len(df_personal)} filas, {len(df_personal.columns)} columnas")
            else:
                if progress_callback:
                    progress_callback("⚠️ Archivo PERSONAL.xlsx no encontrado, solo cruce básico")
            
            # 2. PASO 2: Procesar cruces básicos en paralelo ultra-rápido
            if progress_callback:
                progress_callback("🔥 Iniciando cruces básicos en paralelo...")
            
            with ThreadPoolExecutor(max_workers=2) as executor:
                futures = {executor.submit(process_single_cross_ultra, file_type): file_type 
                          for file_type in self.file_types}
                
                for future in futures:
                    result = future.result()
                    if result:
                        successful_files.append(result)
            
            # 3. PASO 3: Cruce con PERSONAL ultra-optimizado (si está disponible)
            if df_personal is not None and successful_files:
                if progress_callback:
                    progress_callback("👥 Iniciando cruce con PERSONAL ultra-optimizado...")
                
                def process_personal_cross_ultra(file_type):
                    """Procesar cruce con personal ultra-optimizado"""
                    try:
                        # Leer archivo base ya creado
                        base_path = os.path.join('descargas', file_type, f'consolidado_final_{file_type}.xlsx')
                        df_base = pd.read_excel(base_path, engine='openpyxl')
                    
                        if progress_callback:
                            progress_callback(f"👥 Cruzando {file_type} con PERSONAL...")
                        
                        # MERGE ULTRA-OPTIMIZADO con personal
                        # Preparar columnas vectorialmente
                        df_base['OPERADOR_CLEAN'] = df_base['OPERADOR'].str.upper().str.strip()
                        
                        # Merge preservando TODAS las columnas de PERSONAL
                        df_final = pd.merge(
                            df_base,
                            df_personal,
                            left_on='OPERADOR_CLEAN',
                            right_on='APELLIDOS Y NOMBRES',
                            how='left'
                        )
                        
                        # Limpiar columnas duplicadas vectorialmente
                        if 'APELLIDOS Y NOMBRES' in df_final.columns:
                            df_final = df_final.drop(['APELLIDOS Y NOMBRES', 'OPERADOR_CLEAN'], axis=1)
                        else:
                            df_final = df_final.drop('OPERADOR_CLEAN', axis=1)
                        
                        # ESCRITURA ULTRA-RÁPIDA con xlsxwriter
                        output_path_personal = os.path.join('descargas', file_type, f'consolidado_final_{file_type}_personal.xlsx')
                        temp_path_personal = output_path_personal.replace('.xlsx', '_temp_personal.xlsx')
                        
                        workbook_personal = xlsxwriter.Workbook(temp_path_personal, {
                            'constant_memory': True,
                            'nan_inf_to_errors': True,
                            'tmpdir': '.',
                        })
                        worksheet_personal = workbook_personal.add_worksheet('BASE')
                        
                        # Formato de fecha
                        date_format = workbook_personal.add_format({'num_format': 'dd/mm/yyyy'})
                        
                        # Limpieza vectorizada
                        df_final = df_final.replace([np.inf, -np.inf], np.nan)
                        df_final = df_final.fillna('')
                        
                        # Headers
                        for col_idx, col_name in enumerate(df_final.columns):
                            worksheet_personal.write(0, col_idx, col_name)
                        
                        # Identificar columnas de fecha
                        date_cols = [col for col in df_final.columns if 'FECHA' in col.upper()]
                        date_col_indices = [df_final.columns.get_loc(col) for col in date_cols]
                        
                        # Escritura por chunks
                        chunk_size = 6000  # Más pequeño por el archivo más grande
                        total_rows = len(df_final)
                        
                        for start_idx in range(0, total_rows, chunk_size):
                            end_idx = min(start_idx + chunk_size, total_rows)
                            chunk_data = df_final.iloc[start_idx:end_idx].values
                            
                            for row_offset, row_data in enumerate(chunk_data):
                                row_idx = start_idx + row_offset + 1
                                
                                for col_idx, value in enumerate(row_data):
                                    if col_idx in date_col_indices and pd.notna(value) and value != '':
                                        worksheet_personal.write(row_idx, col_idx, value, date_format)
                                    else:
                                        safe_value = str(value) if pd.notna(value) and value != '' else ''
                                        worksheet_personal.write(row_idx, col_idx, safe_value)
                        
                        # Tabla y anchos
                        last_col = xlsxwriter.utility.xl_col_to_name(len(df_final.columns) - 1)
                        table_range = f'A1:{last_col}{len(df_final) + 1}'
                        
                        worksheet_personal.add_table(table_range, {
                            'name': 'BASE',
                            'style': 'Table Style Medium 2',
                            'first_column': False,
                            'last_column': False,
                            'banded_rows': True,
                            'banded_columns': False,
                        })
                        
                        # Anchos optimizados
                        for col_idx in range(len(df_final.columns)):
                            col_name = df_final.columns[col_idx]
                            if df_final[col_name].dtype == 'object':
                                sample_data = df_final[col_name].dropna().astype(str).head(400)
                                if len(sample_data) > 0:
                                    max_len = sample_data.str.len().max()
                                    optimal_width = min(max(max_len, len(col_name)) + 2, 35)
                                else:
                                    optimal_width = len(col_name) + 2
                            else:
                                optimal_width = max(10, len(col_name) + 2)
                            
                            worksheet_personal.set_column(col_idx, col_idx, optimal_width)
                        
                        workbook_personal.close()
                        
                        # Mover archivo final
                        shutil.move(temp_path_personal, output_path_personal)
                        
                        if progress_callback:
                            progress_callback(f"✅ Cruce con PERSONAL completado para {file_type}: {len(df_final.columns)} columnas")
                        
                        return file_type
                    
                    except Exception as e:
                        temp_path_personal = os.path.join('descargas', file_type, f'consolidado_final_{file_type}_temp_personal.xlsx')
                        if os.path.exists(temp_path_personal):
                            os.remove(temp_path_personal)
                        
                        if progress_callback:
                            progress_callback(f"❌ Error en cruce PERSONAL para {file_type}: {str(e)}")
                        return None
                
                # Procesar cruce con PERSONAL en paralelo
                with ThreadPoolExecutor(max_workers=2) as executor:
                    futures_personal = {executor.submit(process_personal_cross_ultra, file_type): file_type 
                                      for file_type in successful_files}
                    
                    for future in futures_personal:
                        future.result()  # Esperar completación
            
            # 4. REPORTE FINAL
            elapsed_time = time.time() - start_time
            if progress_callback:
                if successful_files:
                    progress_callback(f"🎉 Procesamiento de cruces ULTRA-OPTIMIZADO completado!")
                    progress_callback(f"📁 Archivos procesados: {', '.join(successful_files)}")
                    progress_callback(f"⚡ Tiempo total: {elapsed_time:.1f} segundos")
                    progress_callback(f"🚀 Optimización extrema: vectorización + xlsxwriter + threading paralelo")
                    progress_callback(f"💾 Archivos creados: base + personal + CSV para cada tipo")
                else:
                    progress_callback("⚠️ No se pudo procesar ningún cruce")
                
        except Exception as e:
            if progress_callback:
                progress_callback(f"❌ Error en procesamiento ultra-optimizado: {str(e)}")
        
        return successful_files

    def ultra_threaded_format_files(self, progress_callback: Optional[Callable[[str], None]] = None) -> List[str]:
        """
        Método ULTRA-THREADING para formatear archivos Excel.
        Versión MÁS ESTABLE que evita problemas de multiprocessing.
        Usa: vectorización total + xlsxwriter + threading optimizado.
        
        Returns:
            List[str]: Lista de archivos procesados exitosamente
        """
        import pandas as pd
        import xlsxwriter
        from datetime import datetime

        successful_files = []
        
        def ultra_format_single_file_safe(file_type):
            """Formatear archivo con optimizaciones extremas pero estables"""
            try:
                file_path = os.path.join('descargas', file_type, f'consolidado_final_{file_type}_personal.xlsx')
                if not os.path.exists(file_path):
                    if progress_callback:
                        progress_callback(f"⚠️ Archivo no encontrado: {file_path}")
                    return None
                
                if progress_callback:
                    progress_callback(f"🚀 ULTRA-formateando {file_type} (versión estable)...")
                
                # 1. Lectura optimizada
                df = pd.read_excel(file_path, engine='openpyxl')
                
                # 2. VECTORIZACIÓN EXTREMA pero segura
                date_columns = ['FechaExpendiente', 'FechaEtapaAprobacionMasivaFin', 'FechaPre', 'FECHA_ASIGNACION']
                df_processed = df.copy()
                fechapre_converted = 0
                
                for col_name in date_columns:
                    if col_name in df_processed.columns:
                        original_values = df_processed[col_name]
                        mask_not_null = pd.notna(original_values)
                        
                        if mask_not_null.any():
                            try:
                                date_series = pd.to_datetime(original_values[mask_not_null], 
                                                            errors='coerce', 
                                                            dayfirst=True)
                                excel_serial = (date_series - pd.Timestamp('1900-01-01')).dt.days + 2
                                df_processed.loc[mask_not_null, col_name] = excel_serial
                                
                                if col_name == 'FechaPre':
                                    fechapre_converted = excel_serial.notna().sum()
                                
                            except Exception as e:
                                if progress_callback:
                                    progress_callback(f"Warning: Error en conversión de {col_name}: {str(e)}")
                
                if progress_callback:
                    progress_callback(f"🔍 {fechapre_converted} celdas de FechaPre convertidas (método estable)")
                
                # 3. ESCRITURA ULTRA-OPTIMIZADA pero segura
                temp_path = file_path.replace('.xlsx', '_safe_temp.xlsx')
                
                # Configuración segura de xlsxwriter
                workbook = xlsxwriter.Workbook(temp_path, {
                    'constant_memory': True,
                    'nan_inf_to_errors': True,
                    'tmpdir': '.',
                })
                worksheet = workbook.add_worksheet('BASE')
                
                # Formatos
                date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
                number_format = workbook.add_format({'num_format': '#,##0'})
                
                # 4. Escritura de headers
                for col_idx, col_name in enumerate(df_processed.columns):
                    worksheet.write(0, col_idx, col_name)
                
                # 5. LIMPIEZA SEGURA de datos
                df_processed = df_processed.replace([np.inf, -np.inf], np.nan)
                df_processed = df_processed.fillna('')
                
                # 6. Escritura de datos por chunks medianos (balance velocidad/estabilidad)
                chunk_size = 8000
                total_rows = len(df_processed)
                
                for start_idx in range(0, total_rows, chunk_size):
                    end_idx = min(start_idx + chunk_size, total_rows)
                    chunk_data = df_processed.iloc[start_idx:end_idx].values
                    
                    for row_offset, row_data in enumerate(chunk_data):
                        row_idx = start_idx + row_offset + 1
                        
                        for col_idx, value in enumerate(row_data):
                            col_name = df_processed.columns[col_idx]
                            
                            try:
                                if col_name in date_columns and pd.notna(value) and value != '':
                                    worksheet.write(row_idx, col_idx, value, date_format)
                                elif isinstance(value, (int, float)) and col_name not in date_columns:
                                    if pd.notna(value) and not (np.isinf(value) or np.isnan(value)):
                                        worksheet.write(row_idx, col_idx, value, number_format)
                                    else:
                                        worksheet.write(row_idx, col_idx, '')
                                else:
                                    safe_value = str(value) if pd.notna(value) and value != '' else ''
                                    worksheet.write(row_idx, col_idx, safe_value)
                            except Exception as write_error:
                                # En caso de error, escribir como string vacío
                                worksheet.write(row_idx, col_idx, '')
                    
                    # Progreso cada chunk
                    if progress_callback and start_idx % (chunk_size * 3) == 0:
                        progress_pct = (end_idx / total_rows) * 100
                        progress_callback(f"📝 Escribiendo {file_type}: {progress_pct:.1f}%")
                
                # 7. Configuración de anchos optimizada
                for col_idx, col_name in enumerate(df_processed.columns):
                    if df_processed[col_name].dtype == 'object':
                        sample_data = df_processed[col_name].dropna().astype(str).head(800)
                        if len(sample_data) > 0:
                            max_len = sample_data.str.len().max()
                            optimal_width = min(max(max_len, len(col_name)) + 2, 45)
                        else:
                            optimal_width = len(col_name) + 2
                    else:
                        optimal_width = max(12, len(col_name) + 2)
                    
                    worksheet.set_column(col_idx, col_idx, optimal_width)
                
                # 8. Tabla
                last_col = xlsxwriter.utility.xl_col_to_name(len(df_processed.columns) - 1)
                table_range = f'A1:{last_col}{len(df_processed) + 1}'
                
                worksheet.add_table(table_range, {
                    'name': 'BASE',
                    'style': 'Table Style Medium 2',
                    'first_column': False,
                    'last_column': False,
                    'banded_rows': True,
                    'banded_columns': False,
                })
                
                workbook.close()
                
                # Reemplazar archivo
                shutil.move(temp_path, file_path)
                
                # Limpieza de memoria
                del df, df_processed, chunk_data
                gc.collect()
                
                if progress_callback:
                    progress_callback(f"✅ {file_type} ULTRA-formateado exitosamente (versión estable)!")
                
                return file_type
                
            except Exception as e:
                # Limpiar archivo temporal
                temp_path = file_path.replace('.xlsx', '_safe_temp.xlsx')
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                
                if progress_callback:
                    progress_callback(f"❌ Error en ULTRA-formateo estable de {file_type}: {str(e)}")
                return None
        
        try:
            if progress_callback:
                progress_callback("🚀 Iniciando ULTRA-formateo ESTABLE...")
                progress_callback("⚡ Optimización: vectorización + xlsxwriter + threading estable")
            
            # Procesamiento threading estable
            with ThreadPoolExecutor(max_workers=2) as executor:
                futures = {executor.submit(ultra_format_single_file_safe, file_type): file_type 
                          for file_type in ['CCM', 'PRR']}
                
                for future in futures:
                    result = future.result()
                    if result:
                        successful_files.append(result)
            
            if progress_callback:
                if successful_files:
                    progress_callback("🎉 ULTRA-formateo ESTABLE completado!")
                    progress_callback(f"📁 Archivos: {', '.join(successful_files)}")
                    progress_callback("⚡ Velocidad alta + máxima estabilidad garantizada")
                    progress_callback("✅ FechaPre convertida correctamente con método más seguro")
                else:
                    progress_callback("⚠️ No se pudo formatear ningún archivo")
            
        except Exception as e:
            if progress_callback:
                progress_callback(f"❌ Error en ULTRA-formateo estable: {str(e)}")
        
        return successful_files
        
    def optimized_cross_processing(self, progress_callback: Optional[Callable[[str], None]] = None) -> List[str]:
        """
        Método optimizado para procesamiento de cruces usando técnicas avanzadas
        """
        start_time = time.time()
        successful_files = []
        
        if progress_callback:
            progress_callback("🚀 Iniciando procesamiento de cruces optimizado...")
            
        try:
            # 1. Pre-cargar archivo PERSONAL una sola vez y optimizarlo
            personal_path = os.path.join(self.base_dir, 'PERSONAL.xlsx')
            if os.path.exists(personal_path):
                if progress_callback:
                    progress_callback("📊 Cargando y optimizando archivo PERSONAL...")
                
                df_personal = pd.read_excel(personal_path)
                # Optimizar el DataFrame PERSONAL
                df_personal['APELLIDOS Y NOMBRES'] = df_personal['APELLIDOS Y NOMBRES'].str.upper().str.strip()
                
                # NO crear índice aquí para mantener todas las columnas disponibles
            else:
                if progress_callback:
                    progress_callback("⚠️ Archivo PERSONAL.xlsx no encontrado, continuando sin cruce de personal")
                df_personal = None
            
            # 2. Procesar cruces regulares de forma optimizada
            for file_type in self.file_types:
                try:
                    if progress_callback:
                        progress_callback(f"🔄 Procesando cruce optimizado para {file_type}...")
                    
                    # Leer archivos con optimizaciones
                    csv_path = os.path.join('descargas', file_type, f'consolidado_total_{file_type}.csv')
                    xlsx_path = os.path.join(self.base_dir, f'{file_type}-PROCESADO.xlsx')
                    
                    if not os.path.exists(csv_path) or not os.path.exists(xlsx_path):
                        if progress_callback:
                            progress_callback(f"⚠️ Archivos faltantes para {file_type}, saltando...")
                        continue
                    
                    # Lectura optimizada con tipos específicos
                    df_csv = pd.read_csv(csv_path, dtype={'8': str, '9': str, '10': str}, 
                                       low_memory=False)
                    df_xlsx = pd.read_excel(xlsx_path)
                    
                    # Eliminar duplicados antes del merge (más eficiente)
                    df_xlsx_clean = df_xlsx.drop_duplicates(subset=['TRAMITE'], keep='first')
                    
                    if progress_callback:
                        progress_callback(f"🔗 Realizando merge optimizado para {file_type}...")
                    
                    # Merge optimizado - paso 1: cruce con datos procesados
                    df_merged = pd.merge(
                        df_csv,
                        df_xlsx_clean[['TRAMITE', 'OPERADOR', 'FECHA_ASIGNACION']],
                        left_on='NumeroTramite',
                        right_on='TRAMITE',
                        how='left'
                    ).drop('TRAMITE', axis=1)
                    
                    # Guardar archivo consolidado sin personal primero
                    output_path_base = os.path.join('descargas', file_type, f'consolidado_final_{file_type}.xlsx')
                    
                    with pd.ExcelWriter(output_path_base, engine='openpyxl', 
                                      date_format='dd/mm/yyyy') as writer:
                        df_merged.to_excel(writer, index=False)
                        
                        # Formateo optimizado solo para columnas de fecha
                        worksheet = writer.sheets['Sheet1']
                        date_cols = [col for col in df_merged.columns if 'FECHA' in col.upper()]
                        
                        for col_name in date_cols:
                            col_idx = df_merged.columns.get_loc(col_name) + 1
                            col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + col_idx//26) + chr(64 + col_idx%26)
                            
                            # Aplicar formato a toda la columna de una vez
                            for cell in worksheet[f"{col_letter}2:{col_letter}{len(df_merged) + 1}"]:
                                if isinstance(cell, tuple):
                                    for c in cell:
                                        c.number_format = 'dd/mm/yyyy'
                                else:
                                    cell.number_format = 'dd/mm/yyyy'
                    
                    # Cruce con PERSONAL si está disponible (paso 2)
                    if df_personal is not None:
                        if progress_callback:
                            progress_callback(f"👥 Realizando cruce con PERSONAL para {file_type}...")
                        
                        # Preparar columna para merge más eficiente
                        df_merged['OPERADOR_CLEAN'] = df_merged['OPERADOR'].str.upper().str.strip()
                        
                        # Merge con PERSONAL manteniendo TODAS las columnas
                        df_final = pd.merge(
                            df_merged,
                            df_personal,
                            left_on='OPERADOR_CLEAN',
                            right_on='APELLIDOS Y NOMBRES',
                            how='left'
                        )
                        
                        # Limpiar columnas duplicadas
                        if 'APELLIDOS Y NOMBRES' in df_final.columns and 'OPERADOR' in df_final.columns:
                            df_final = df_final.drop('APELLIDOS Y NOMBRES', axis=1)
                        
                        df_final.drop('OPERADOR_CLEAN', axis=1, inplace=True)
                        
                        # Guardar archivo consolidado CON personal
                        output_path_personal = os.path.join('descargas', file_type, f'consolidado_final_{file_type}_personal.xlsx')
                        
                        with pd.ExcelWriter(output_path_personal, engine='openpyxl', 
                                          date_format='dd/mm/yyyy') as writer:
                            df_final.to_excel(writer, index=False)
                            
                            # Formateo optimizado solo para columnas de fecha
                            worksheet = writer.sheets['Sheet1']
                            date_cols = [col for col in df_final.columns if 'FECHA' in col.upper()]
                            
                            for col_name in date_cols:
                                col_idx = df_final.columns.get_loc(col_name) + 1
                                col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + col_idx//26) + chr(65 + col_idx%26)
                                
                    successful_files.append(file_type)
                    
                except Exception as e:
                    if progress_callback:
                        progress_callback(f"❌ Error regenerando {file_type}: {str(e)}")
                    continue
            
            if progress_callback:
                if successful_files:
                    progress_callback(f"🎉 Regeneración completada. Archivos: {', '.join(successful_files)}")
                else:
                    progress_callback("⚠️ No se pudo regenerar ningún archivo")
                
        except Exception as e:
            if progress_callback:
                progress_callback(f"❌ Error en regeneración: {str(e)}")
        
        return successful_files
        