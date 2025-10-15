import tkinter as tk
from tkinter import scrolledtext, filedialog, messagebox
import ttkbootstrap as tb
from ttkbootstrap.constants import *
import threading
import queue
import pandas as pd
import os
import sys
import requests
import json
from pathlib import Path
from typing import Optional
from requests_ntlm import HttpNtlmAuth
from descarga import descargar_y_consolidar, BASE_URL
from utils_logging import logger, set_gui_mode
from datetime import datetime, date
from excel_processor import ExcelProcessor
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from credentials_manager import credentials_manager
from structure_validator import validate_project_structure
import time


def _resource_path(relative_path: str) -> Path:
    """
    Devuelve la ruta absoluta a un recurso, compatible con PyInstaller.
    """
    base_path = getattr(sys, "_MEIPASS", Path(__file__).resolve().parent.parent)
    return Path(base_path) / relative_path


class MainWindow(tb.Window):
    def __init__(self, run_validation=True):
        super().__init__(themename="superhero")
        set_gui_mode(True)
        self._icon_photo = None
        self._set_window_icon()
        
        if not run_validation:
            # Si la validación no se ejecuta, es porque estamos mostrando
            # un error crítico desde main.py. No se debe construir la GUI.
            return

        # Configure window style and theme
        self.title("Reporteador - Sistema de Descarga y Procesamiento")
        self.geometry("1100x800")
        self.minsize(900, 700)
        
        self.process_thread = None
        self.queue = queue.Queue()
        self.is_running = False

        # Ya no se llama a self.validate_project_structure() aquí

        # Load module selection from environment
        self.load_module_selection()

        self.create_widgets()
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.schedule_file = "schedule_config.json"
        self.schedule_lock = threading.Lock()
        self.last_run_date = None
        self.scheduled_time = None
        
        self.load_schedule()
        self.start_scheduler_thread()

    def _set_window_icon(self):
        """Configura el icono de la ventana en runtime."""
        try:
            icon_path = _resource_path('assets/reporteador.ico')
            if icon_path.exists():
                try:
                    self.iconbitmap(str(icon_path))
                except Exception as exc:
                    logger.warning(f"No se pudo aplicar icono .ico: {exc}")

            png_path = _resource_path('assets/reporteador.png')
            if png_path.exists():
                try:
                    self._icon_photo = tk.PhotoImage(file=str(png_path))
                    self.iconphoto(False, self._icon_photo)
                except Exception as exc:
                    logger.warning(f"No se pudo establecer icono PNG: {exc}")
        except Exception as exc:
            logger.warning(f"No se pudo configurar el icono de la aplicación: {exc}")

    def load_module_selection(self):
        """Load module selection state from environment variables"""
        try:
            selected_modules = json.loads(os.getenv('SELECTED_MODULES', '{}'))
            self.module_vars = {
                "CCM": tk.BooleanVar(value=selected_modules.get("CCM", True)),
                "PRR": tk.BooleanVar(value=selected_modules.get("PRR", True)),
                
            }
        except json.JSONDecodeError:
            self.module_vars = {
                "CCM": tk.BooleanVar(value=True),
                "PRR": tk.BooleanVar(value=True),
                
            }

    def save_module_selection(self):
        """Save current module selection to environment variables"""
        selected_modules = {name: var.get() for name, var in self.module_vars.items()}
        os.environ['SELECTED_MODULES'] = json.dumps(selected_modules)

    def validate_project_structure(self):
        """
        Valida y crea automáticamente la estructura de carpetas y archivos necesarios
        """
        def validation_log(message):
            """Log específico para validación - solo errores críticos se muestran al usuario"""
            logger.info(message)
            # Solo mostrar errores críticos en un diálogo
            if "❌" in message and ("Error" in message or "Faltante" in message):
                print(f"Validación: {message}")  # Para debug
        
        try:
            # Ejecutar validación con creación automática
            success, report = validate_project_structure(
                base_path=None,  # Usar directorio actual
                log_callback=validation_log,
                auto_create=True
            )
            
            # Solo mostrar diálogo si hay errores críticos que requieren atención del usuario
            critical_errors = [error for error in report.get('errors', []) 
                             if 'permisos' in error.lower() or 'módulos python faltantes' in error.lower()]
            
            if critical_errors:
                error_message = "Se encontraron problemas críticos que requieren atención:\n\n"
                error_message += "\n".join(f"• {error}" for error in critical_errors)
                error_message += "\n\nLa aplicación puede no funcionar correctamente."
                
                messagebox.showwarning(
                    "Problemas de Estructura Detectados",
                    error_message
                )
            
            # Mostrar advertencias importantes sobre archivos opcionales
            important_warnings = [warning for warning in report.get('warnings', [])
                                if 'PERSONAL.xlsx' in warning or 'NTLM' in warning]
            
            if important_warnings and len(important_warnings) == 1 and 'PERSONAL.xlsx' in important_warnings[0]:
                # Solo mostrar advertencia sobre PERSONAL.xlsx si es la única advertencia importante
                messagebox.showinfo(
                    "Información Importante",
                    "⚠️ Archivo PERSONAL.xlsx no encontrado en ASIGNACIONES/\n\n"
                    "Este archivo es necesario para los procesos de cruce de datos.\n"
                    "Puedes agregarlo manualmente cuando lo tengas disponible."
                )
            
            # Log resumen para desarrollador
            created_count = len(report.get('directories', {}).get('created', [])) + len(report.get('files', {}).get('created', []))
            if created_count > 0:
                logger.info(f"✅ Estructura validada - {created_count} elementos creados automáticamente")
            else:
                logger.info("✅ Estructura del proyecto ya estaba completa")
                
        except Exception as e:
            logger.error(f"Error durante validación de estructura: {str(e)}")
            # Solo mostrar error si es algo realmente crítico
            if "Permission" in str(e) or "Access" in str(e):
                messagebox.showerror(
                    "Error de Permisos",
                    f"No se pudo validar/crear la estructura del proyecto:\n{str(e)}\n\n"
                    "Ejecuta la aplicación como administrador o verifica los permisos."
                )

    def create_widgets(self):
        # Create main frame with proper padding
        self.main_frame = tb.Frame(self, padding="20")
        self.main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure main window grid
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # Create notebook with improved styling
        self.notebook = tb.Notebook(self.main_frame, bootstyle="primary")
        self.notebook.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5, pady=5)

        # Create tabs with consistent styling
        self.automated_tab = tb.Frame(self.notebook, padding="10")
        self.descarga_tab = tb.Frame(self.notebook, padding="10")
        self.procesamiento_tab = tb.Frame(self.notebook, padding="10")
        self.settings_tab = tb.Frame(self.notebook, padding="10")
        
        self.notebook.add(self.automated_tab, text='Proceso Automatizado')
        self.notebook.add(self.descarga_tab, text='Descarga')
        self.notebook.add(self.procesamiento_tab, text='Procesamiento')
        self.notebook.add(self.settings_tab, text='Configuración')

        self.setup_download_tab()
        self.setup_processing_tab()
        self.setup_settings_tab()
        self.setup_automated_tab()

        # Configure main frame expansion
        self.main_frame.grid_columnconfigure(0, weight=1)
        self.main_frame.grid_rowconfigure(0, weight=1)

    def setup_download_tab(self):
        # Download Options with improved layout
        options_frame = tb.LabelFrame(self.descarga_tab, text="Opciones de Descarga", padding="10")
        options_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), padx=10, pady=5)
        
        # Error log viewer
        error_log_frame = tb.LabelFrame(self.descarga_tab, text="Registro de Errores", padding="10")
        error_log_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), padx=10, pady=5)
        error_log_frame.grid_columnconfigure(0, weight=1)
        error_log_frame.grid_rowconfigure(0, weight=1)
        
        self.error_log = scrolledtext.ScrolledText(error_log_frame, height=5, wrap=tk.WORD)
        self.error_log.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5, pady=5)
        self.error_log.tag_configure("error", foreground="red")
        
        # Download mode options with better spacing
        mode_frame = tb.Frame(options_frame)
        mode_frame.pack(fill=tk.X, padx=10, pady=10)
        
        self.download_option = tk.StringVar(value="all")
        tb.Radiobutton(mode_frame, text="Descargar Todo", variable=self.download_option, value="all", bootstyle="primary").pack(side=tk.LEFT, padx=10)
        tb.Radiobutton(mode_frame, text="Descargar Faltantes", variable=self.download_option, value="missing", bootstyle="primary").pack(side=tk.LEFT, padx=10)
        tb.Radiobutton(mode_frame, text="Solo Consolidar", variable=self.download_option, value="consolidate", bootstyle="primary").pack(side=tk.LEFT, padx=10)
        
        # Module selection with improved visual grouping
        module_frame = tb.LabelFrame(options_frame, text="Módulos", padding="5")
        module_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        for module, var in self.module_vars.items():
            cb = tb.Checkbutton(module_frame, text=module, variable=var, command=self.save_module_selection, bootstyle="primary")
            cb.pack(side=tk.LEFT, padx=15, pady=5)

        # Log area with better visual hierarchy
        log_frame = tb.LabelFrame(self.descarga_tab, text="Registro de Operaciones", padding="10")
        log_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), padx=10, pady=5)
        log_frame.grid_columnconfigure(0, weight=1)
        log_frame.grid_rowconfigure(0, weight=1)

        self.log_area = scrolledtext.ScrolledText(log_frame, height=15, wrap=tk.WORD)
        self.log_area.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5, pady=5)

        # Progress section with improved layout
        progress_frame = tb.LabelFrame(self.descarga_tab, text="Progreso", padding="10")
        progress_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), padx=10, pady=5)
        progress_frame.grid_columnconfigure(1, weight=1)
        
        self.progress_label = tb.Label(progress_frame, text="Esperando inicio...")
        self.progress_label.grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 5))
        
        self.speed_label = tb.Label(progress_frame, text="")
        self.speed_label.grid(row=1, column=0, sticky=tk.W)
        
        self.eta_label = tb.Label(progress_frame, text="")
        self.eta_label.grid(row=1, column=1, sticky=tk.E)
        
        self.progress_bar = tb.Progressbar(progress_frame, length=300, mode='determinate', bootstyle="success-striped")
        self.progress_bar.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(5, 0))

        # Control buttons with consistent spacing
        button_frame = tb.Frame(self.descarga_tab)
        button_frame.grid(row=3, column=0, columnspan=2, pady=15)

        self.start_button = tb.Button(button_frame, text="Iniciar Proceso", command=self.start_process, bootstyle="success")
        self.start_button.grid(row=0, column=0, padx=10)

        self.pause_button = tb.Button(button_frame, text="Pausar", command=self.toggle_pause, state=tk.DISABLED, bootstyle="warning")
        self.pause_button.grid(row=0, column=1, padx=10)

        self.cancel_button = tb.Button(button_frame, text="Cancelar", command=self.cancel_process, state=tk.DISABLED, bootstyle="danger")
        self.cancel_button.grid(row=0, column=2, padx=10)

        # Configure tab expansion
        self.descarga_tab.grid_columnconfigure(0, weight=1)
        self.descarga_tab.grid_rowconfigure(1, weight=1)

    def toggle_pause(self):
        if hasattr(self, 'download_manager'):
            if self.download_manager.is_paused:
                self.download_manager.set_paused(False)
                self.pause_button.configure(text="Pausar")
                self.log_message("Descarga reanudada")
            else:
                self.download_manager.set_paused(True)
                self.pause_button.configure(text="Reanudar")
                self.log_message("Descarga pausada")

    def update_download_stats(self, progress_data):
        progress = progress_data.get('progress', 0)
        speed = progress_data.get('speed', 0)
        estimated_time = progress_data.get('estimated_time', 0)
        
        # Update progress bar
        self.progress_bar["value"] = progress
        
        # Update speed label
        speed_mb = speed / (1024 * 1024)  # Convert to MB/s
        self.speed_label.configure(text=f"Velocidad: {speed_mb:.2f} MB/s")
        
        # Format estimated time
        if estimated_time > 0:
            minutes = int(estimated_time // 60)
            seconds = int(estimated_time % 60)
            time_str = f"Tiempo restante: {minutes}m {seconds}s"
        else:
            time_str = "Calculando tiempo..."
        
        # Update time label
        if not hasattr(self, 'time_label'):
            self.time_label = tb.Label(self.status_frame, text=time_str)
            self.time_label.pack(side=tk.LEFT, padx=5)
        else:
            self.time_label.configure(text=time_str)
        
        # Update progress label
        phase = "Descarga" if progress <= 80 else "Consolidación"
        self.progress_label.configure(text=f"{phase}: {progress:.1f}%")

    def start_process(self):
        if not self.is_running:
            # La validación de carpetas ahora se hace al inicio.
            # Solo validamos la conexión de red y la selección.
            if not self.check_network_status():
                messagebox.showerror("Error de Red", "No hay conexión a Internet.")
                return

            selected_modules = {name: var.get() for name, var in self.module_vars.items()}
            if not any(selected_modules.values()):
                messagebox.showerror("Error", "Por favor, seleccione al menos un módulo.")
                return

            if not credentials_manager.has_credentials():
                messagebox.showerror(
                    "Credenciales NTLM requeridas",
                    "Configura las credenciales NTLM en la pestaña de configuración antes de iniciar el proceso."
                )
                return

            self.is_running = True
            self.start_button.configure(state=tk.DISABLED)
            self.cancel_button.configure(state=tk.NORMAL)
            self.pause_button.configure(state=tk.NORMAL)
            self.progress_bar["value"] = 0
            self.log_area.delete(1.0, tk.END)
            self.log_message("Iniciando proceso de descarga...")
            saved_user, _ = credentials_manager.get_credentials()
            if saved_user:
                self.log_message(f"Usando credenciales NTLM guardadas para: {saved_user}")
            else:
                self.log_message("No hay credenciales NTLM guardadas. Configúralas en la pestaña de configuración.")
            self.update_network_status()

            # Update environment variables with current settings
            os.environ['DOWNLOAD_MAX_WORKERS'] = self.workers_var.get()
            os.environ['DOWNLOAD_DELAY'] = self.delay_var.get()
            os.environ['CHUNK_SIZE'] = self.chunk_var.get()
            os.environ['DIRECT_DOWNLOAD'] = str(self.direct_download_var.get()).lower()

            def log_callback(message):
                self.queue.put(("log", message))

            def error_callback(file_path, error_message):
                full_error = f"{error_message} [{file_path}]"
                self.queue.put(("error", full_error))

            def progress_callback(progress):
                self.queue.put(("progress", progress))

            def run_process():
                try:
                    download_option = self.download_option.get()
                    selected_modules = {name: var.get() for name, var in self.module_vars.items()}
                    descargar_y_consolidar(
                        download_option=download_option,
                        progress_callback=progress_callback,
                        log_callback=log_callback,
                        error_callback=error_callback,
                        selected_modules=selected_modules,
                        overwrite=True
                    )
                except Exception as e:
                    self.queue.put(("error", str(e)))
                finally:
                    self.queue.put(("finished", None))

            self.process_thread = threading.Thread(target=run_process)
            self.process_thread.start()
            self.check_queue()

    def check_queue(self):
        try:
            while True:
                item = self.queue.get_nowait()
                message_type = item[0]
                message_data = item[1]
                
                if message_type == "log":
                    self.log_message(message_data)
                elif message_type == "error":
                    self.log_message(message_data, error=True)
                elif message_type == "progress":
                    self.update_progress(message_data)
                elif message_type == "process_log":
                    self.log_message(message_data, process=True)
                elif message_type == "process_error":
                    self.log_message(message_data, error=True, process=True)
                elif message_type == "process_progress":
                    self.process_progress_bar["value"] = message_data
                elif message_type == "process_finished":
                    self.process_finished()
                    break
        except queue.Empty:
            pass
        
        if self.is_running:
            self.after(100, self.check_queue)

    def _toggle_chunk_settings(self):
        """Toggle chunk-related settings based on direct download option"""
        is_direct = self.direct_download_var.get()
        chunk_state = 'disabled' if is_direct else 'normal'
        self.chunk_var.configure(state=chunk_state)

    def update_progress(self, value):
        """Actualiza la barra de progreso y el estado de manera estrictamente creciente y fluida"""
        if not isinstance(value, (int, float)) or value < 0:
            return

        # Mantener el progreso anterior para evitar retrocesos
        if not hasattr(self, '_last_progress'):
            self._last_progress = 0
        # Solo permitir que el progreso avance o se mantenga
        progress = max(self._last_progress, min(max(value, 0), 100))
        self._last_progress = progress
        self.progress_bar["value"] = progress

        # Determinar la fase actual
        if progress == 0:
            status = "Iniciando descarga..."
        elif progress < 80:
            status = f"Descarga en progreso: {progress:.1f}%"
        else:
            status = f"Consolidación: {progress:.1f}%"

        self.progress_label.configure(text=status)
        self.progress_bar.update_idletasks()

    def log_message(self, message, error=False, process=False):
        timestamp = datetime.now().strftime("[%H:%M:%S]")
        formatted_message = f"{timestamp} {message}\n"
        
        log_area = self.process_log_area if process else self.log_area
        log_area.insert(tk.END, formatted_message)
        
        # Tag the entire line for coloring
        line_start = "end-2c linestart"
        line_end = "end-1c"
        
        if error:
            log_area.tag_add("error", line_start, line_end)
            log_area.tag_configure("error", foreground="red")
            # Add error to error log viewer
            self.error_log.insert(tk.END, formatted_message)
            self.error_log.tag_add("error", "end-2c linestart", "end-1c")
            self.error_log.see(tk.END)
        else:
            log_area.tag_add("info", line_start, line_end)
            log_area.tag_configure("info", foreground="blue")
            
        log_area.see(tk.END)

    def check_network_status(self):
        try:
            response = requests.get("http://www.google.com", timeout=5)
            return response.status_code == 200
        except requests.RequestException:
            return False

    def update_network_status(self):
        if self.is_running:
            is_connected = self.check_network_status()
            if not is_connected:
                self.log_message("Error de conectividad de red detectado", error=True)
                self.progress_label.configure(text="Sin conexión", foreground="red")
            else:
                self.progress_label.configure(foreground="black")
            self.after(30000, self.update_network_status)  # Check every 30 seconds

    def cancel_process(self):
        if self.is_running:
            self.is_running = False
            self.log_message("Cancelando proceso...")
            self.process_finished()

    def process_finished(self):
        self.is_running = False
        self._enable_all_buttons()
        self.cancel_button.configure(state=tk.DISABLED)
        self.pause_button.configure(state=tk.DISABLED)
        self.progress_label.configure(text="Proceso finalizado")

    def setup_processing_tab(self):
        # Options frame with improved layout
        options_frame = tb.LabelFrame(self.procesamiento_tab, text="Opciones de Procesamiento", padding="10")
        options_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), padx=10, pady=5)
        options_frame.grid_columnconfigure(0, weight=1)

        # Progress section with better visual feedback
        self.process_progress_frame = tb.LabelFrame(self.procesamiento_tab, text="Estado del Proceso", padding="10")
        self.process_progress_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), padx=10, pady=5)
        self.process_progress_frame.grid_columnconfigure(0, weight=1)
        
        self.process_progress_label = tb.Label(self.process_progress_frame, text="Esperando inicio...")
        self.process_progress_label.grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        
        self.process_progress_bar = tb.Progressbar(self.process_progress_frame, mode='determinate', bootstyle="info-striped")
        self.process_progress_bar.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 5))

        # Log area with improved visual hierarchy
        process_log_frame = tb.LabelFrame(self.procesamiento_tab, text="Registro de Procesamiento", padding="10")
        process_log_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), padx=10, pady=5)
        process_log_frame.grid_columnconfigure(0, weight=1)
        process_log_frame.grid_rowconfigure(0, weight=1)

        self.process_log_area = scrolledtext.ScrolledText(process_log_frame, height=15, wrap=tk.WORD)
        self.process_log_area.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5, pady=5)

        # Control buttons with consistent styling - Solo botones necesarios
        process_button_frame = tb.Frame(self.procesamiento_tab)
        process_button_frame.grid(row=3, column=0, columnspan=2, pady=15)

        self.process_button = tb.Button(process_button_frame, text="Procesar Archivos", 
                                       command=self.start_processing, bootstyle="primary")
        self.process_button.grid(row=0, column=0, padx=10)

        self.ultra_cross_button = tb.Button(process_button_frame, text="Cruce",
                                           command=self.start_ultra_fast_cross_processing, bootstyle="primary")
        self.ultra_cross_button.grid(row=0, column=1, padx=10)

        self.max_perf_format_button = tb.Button(process_button_frame, text="Formateo",
                                               command=self.start_max_performance_format_processing, bootstyle="primary")
        self.max_perf_format_button.grid(row=0, column=2, padx=10)

        self.optimize_button = tb.Button(process_button_frame, text="Conversion",
                                        command=self.open_optimize_dialog, bootstyle="info")
        self.optimize_button.grid(row=0, column=3, padx=10)


        # Configure grid weights for proper expansion
        self.procesamiento_tab.grid_columnconfigure(0, weight=1)
        self.procesamiento_tab.grid_rowconfigure(1, weight=1)

    def start_format_processing(self):
        if not self.is_running:
            self.is_running = True
            self._disable_all_buttons()
            self.process_progress_bar["value"] = 0
            self.process_log_area.delete(1.0, tk.END)
            self.log_message("Iniciando formateo de archivos y fechas...", process=True)

            def run_process():
                try:
                    processor = ExcelProcessor()
                    total_files = len(['CCM', 'PRR'])
                    for file_index, file_type in enumerate(['CCM', 'PRR']):
                        file_path = os.path.join('descargas', file_type, f'consolidado_final_{file_type}_personal.xlsx')
                        if os.path.exists(file_path):
                            try:
                                self.queue.put(("process_log", f"Formateando archivo {file_type}..."))
                                
                                # Read Excel file into pandas DataFrame
                                df = pd.read_excel(file_path)
                                
                                # Define date columns
                                date_columns = ['FechaExpendiente', 'FechaEtapaAprobacionMasivaFin', 'FechaPre', 'FECHA_ASIGNACION']
                                
                                # Create a new workbook
                                wb = openpyxl.Workbook()
                                ws = wb.active
                                
                                # Write headers
                                for col_idx, col_name in enumerate(df.columns, 1):
                                    ws.cell(row=1, column=col_idx, value=col_name)
                                
                                # Write data and format dates and numbers
                                for row_idx, row in enumerate(df.itertuples(index=False), 2):
                                    for col_idx, value in enumerate(row, 1):
                                        cell = ws.cell(row=row_idx, column=col_idx)
                                        col_name = df.columns[col_idx - 1]
                                        
                                        if col_name in date_columns:
                                            try:
                                                # Try to parse the date value
                                                if pd.notna(value):
                                                    if isinstance(value, (datetime, pd.Timestamp)):
                                                        date_value = value
                                                    else:
                                                        # Try parsing string dates
                                                        try:
                                                            date_value = pd.to_datetime(str(value), dayfirst=True)
                                                        except:
                                                            try:
                                                                date_value = pd.to_datetime(str(value))
                                                            except:
                                                                continue
                                                    
                                                    # Convert to Excel serial number
                                                    excel_date = date_value.toordinal() - datetime(1900, 1, 1).toordinal() + 2
                                                    cell.value = excel_date
                                                    cell.number_format = 'dd/mm/yyyy'
                                            except Exception as e:
                                                self.queue.put(("process_log", f"Warning: Could not format date in {col_name}: {str(e)}"))
                                                cell.value = value
                                        elif isinstance(value, (int, float)):
                                            cell.value = value
                                            cell.number_format = '#,##0'
                                        else:
                                            cell.value = value
                                
                                # Auto-adjust column widths
                                for column in ws.columns:
                                    max_length = 0
                                    column_letter = get_column_letter(column[0].column)
                                    for cell in column:
                                        try:
                                            if len(str(cell.value)) > max_length:
                                                max_length = len(str(cell.value))
                                        except:
                                            pass
                                    adjusted_width = (max_length + 2)
                                    ws.column_dimensions[column_letter].width = adjusted_width
                                
                                # Convert the data range to a table named 'BASE'
                                tab = Table(displayName="BASE", ref=f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}")
                                style = TableStyleInfo(
                                    name="TableStyleMedium2",
                                    showFirstColumn=False,
                                    showLastColumn=False,
                                    showRowStripes=True,
                                    showColumnStripes=False
                                )
                                tab.tableStyleInfo = style
                                ws.add_table(tab)
                                
                                # Save the workbook
                                wb.save(file_path)
                                
                                # Update progress
                                progress = ((file_index + 1) / total_files) * 100
                                self.queue.put(("process_progress", progress))
                                self.queue.put(("process_log", f"Archivo {file_type} formateado exitosamente."))
                                
                            except Exception as e:
                                self.queue.put(("process_error", f"Error al procesar {file_type}: {str(e)}"))
                                continue

                except Exception as e:
                    self.queue.put(("process_error", str(e)))
                finally:
                    self.queue.put(("process_finished", None))

            self.process_thread = threading.Thread(target=run_process)
            self.process_thread.start()
            self.check_queue()

    def open_optimize_dialog(self):
        """Abre el diálogo de optimización para Streamlit"""
        from gui_optimize_button import OptimizeFilesDialog
        dialog = OptimizeFilesDialog(self)

    def start_processing(self):
        if not self.is_running:
            # La validación de archivos ahora se hace al inicio.
            # Podemos asumir que los archivos base existen si la app inició.
            self.is_running = True
            self._disable_all_buttons()
            self.process_progress_bar["value"] = 0
            self.process_log_area.delete(1.0, tk.END)
            self.log_message("Iniciando procesamiento de archivos...", process=True)

            def run_process():
                try:
                    self.process_excel_files()
                except Exception as e:
                    self.queue.put(("process_error", str(e)))
                finally:
                    self.queue.put(("process_finished", None))

            self.process_thread = threading.Thread(target=run_process)
            self.process_thread.start()
            self.check_queue()

    def start_cross_processing(self):
        if not self.is_running:
            self.is_running = True
            self._disable_all_buttons()
            self.process_progress_bar["value"] = 0
            self.process_log_area.delete(1.0, tk.END)
            self.log_message("Iniciando procesamiento de cruces...", process=True)

            def run_process():
                try:
                    processor = ExcelProcessor()
                    successful_files = processor.process_cross_files(
                        lambda msg: self.queue.put(("process_log", msg))
                    )
                    if not successful_files:
                        self.queue.put(("process_error", "No se procesó ningún cruce correctamente"))
                except Exception as e:
                    self.queue.put(("process_error", str(e)))
                finally:
                    self.queue.put(("process_finished", None))

            self.process_thread = threading.Thread(target=run_process)
            self.process_thread.start()
            self.check_queue()

    def process_calidades_files(self):
        processor = ExcelProcessor()
        
        def progress_callback(message):
            self.queue.put(("process_log", message))
            if "Successfully processed" in message:
                self.queue.put(("process_progress", 100))
        
        try:
            successful_files = processor.process_all_calidades_files(progress_callback)
            if not successful_files:
                self.queue.put(("process_error", "No se procesó ningún archivo CALIDADES correctamente"))
            else:
                self.queue.put(("process_log", f"Procesamiento de CALIDADES completado. Archivos procesados: {', '.join(successful_files)}"))
        except Exception as e:
            self.queue.put(("process_error", f"Error durante el procesamiento de CALIDADES: {str(e)}"))

    def process_excel_files(self):
        processor = ExcelProcessor()
        
        def progress_callback(message):
            self.queue.put(("process_log", message))
            
            # Update progress based on completion status
            if "processed" in message.lower():
                file_type = message.split()[2] if len(message.split()) > 2 else ""
                if file_type in processor.file_types:
                    progress = (processor.file_types.index(file_type) + 1) / len(processor.file_types) * 100
                    self.queue.put(("process_progress", progress))
        
        try:
            successful_files = processor.process_all_files(progress_callback)
            if not successful_files:
                self.queue.put(("process_error", "No se procesó ningún archivo correctamente"))
            else:
                self.queue.put(("process_log", f"Procesamiento completado. Archivos procesados: {', '.join(successful_files)}"))
        except Exception as e:
            self.queue.put(("process_error", f"Error durante el procesamiento: {str(e)}"))

    def setup_settings_tab(self):
        # Settings frame
        settings_frame = tb.Frame(self.settings_tab, padding="10")
        settings_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=10, pady=5)
        settings_frame.grid_columnconfigure(1, weight=1)

        # Advanced Settings with better organization and recommendations
        advanced_frame = tb.LabelFrame(settings_frame, text="Configuración de Velocidad de Descarga", padding="10")
        advanced_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        advanced_frame.grid_columnconfigure(1, weight=1)

        # Workers Setting with recommendations
        workers_frame = tb.Frame(advanced_frame)
        workers_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        tb.Label(workers_frame, text="Número de trabajadores:").pack(side=tk.LEFT, padx=5)
        self.workers_var = tb.Entry(workers_frame, width=5)
        self.workers_var.pack(side=tk.LEFT, padx=5)
        self.workers_var.insert(0, os.getenv('DOWNLOAD_MAX_WORKERS', '5'))
        tb.Label(workers_frame, text="(Procesos paralelos de descarga)").pack(side=tk.LEFT, padx=5)
        tb.Label(workers_frame, 
                  text="Recomendado: 1=Lento, 3=Medio, 5=Rápido, 8=Ultra").pack(side=tk.LEFT, padx=5)

        # Delay Setting with recommendations
        delay_frame = tb.Frame(advanced_frame)
        delay_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        tb.Label(delay_frame, text="Retardo de descarga (s):").pack(side=tk.LEFT, padx=5)
        self.delay_var = tb.Entry(delay_frame, width=5)
        self.delay_var.pack(side=tk.LEFT, padx=5)
        self.delay_var.insert(0, os.getenv('DOWNLOAD_DELAY', '1'))
        tb.Label(delay_frame, text="(Pausa entre descargas)").pack(side=tk.LEFT, padx=5)
        tb.Label(delay_frame, 
                  text="Recomendado: 2.0=Lento, 1.0=Medio, 0.5=Rápido, 0.1=Ultra").pack(side=tk.LEFT, padx=5)

        # Chunk Size Setting with recommendations
        chunk_frame = tb.Frame(advanced_frame)
        chunk_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        tb.Label(chunk_frame, text="Tamaño de chunk (bytes):").pack(side=tk.LEFT, padx=5)
        self.chunk_var = tb.Entry(chunk_frame, width=8)
        self.chunk_var.pack(side=tk.LEFT, padx=5)
        self.chunk_var.insert(0, os.getenv('CHUNK_SIZE', '8192'))
        tb.Label(chunk_frame, text="(Tamaño de bloque de datos)").pack(side=tk.LEFT, padx=5)
        tb.Label(chunk_frame, 
                  text="Recomendado: 4096=Lento, 8192=Medio, 16384=Rápido").pack(side=tk.LEFT, padx=5)

        # Direct Download Option
        direct_download_frame = tb.Frame(advanced_frame)
        direct_download_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        self.direct_download_var = tk.BooleanVar(value=os.getenv('DIRECT_DOWNLOAD', 'true').lower() == 'true')
        direct_download_cb = tb.Checkbutton(direct_download_frame, 
                                           text="Descarga directa (sin chunks)",
                                           variable=self.direct_download_var,
                                           command=self._toggle_chunk_settings,
                                           bootstyle="primary")
        direct_download_cb.pack(side=tk.LEFT, padx=5)
        tb.Label(direct_download_frame, 
                  text="(Recomendado para archivos pequeños o conexiones estables)").pack(side=tk.LEFT, padx=5)

        # NTLM Credentials
        ntlm_frame = tb.LabelFrame(settings_frame, text="Credenciales NTLM", padding="10")
        ntlm_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        ntlm_frame.grid_columnconfigure(1, weight=1)

        # NTLM User
        tb.Label(ntlm_frame, text="Usuario NTLM:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        
        saved_user, saved_pass = credentials_manager.get_credentials()
        
        self.ntlm_user_var = tk.StringVar(value=saved_user or '')
        ntlm_user_entry = tb.Entry(ntlm_frame, textvariable=self.ntlm_user_var, width=40)
        ntlm_user_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)

        # NTLM Password
        tb.Label(ntlm_frame, text="Contraseña NTLM:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.ntlm_pass_var = tk.StringVar(value=saved_pass or '')
        ntlm_pass_entry = tb.Entry(ntlm_frame, textvariable=self.ntlm_pass_var, show="*", width=40)
        ntlm_pass_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)
        
        # Status label with current account
        self.ntlm_status_label = tb.Label(
            ntlm_frame,
            text=self._format_ntlm_status(saved_user),
            bootstyle="secondary"
        )
        self.ntlm_status_label.grid(row=2, column=0, columnspan=2, sticky=tk.W, padx=5, pady=(5, 0))
        
        # Buttons frame for NTLM actions
        ntlm_buttons_frame = tb.Frame(ntlm_frame)
        ntlm_buttons_frame.grid(row=3, column=1, sticky=tk.E, padx=5, pady=10)
        
        # Save NTLM Credentials Button
        save_ntlm_button = tb.Button(ntlm_buttons_frame, text="Guardar Credenciales", command=self.save_ntlm_credentials, bootstyle="success")
        save_ntlm_button.pack(side=tk.LEFT, padx=(0, 5))
        
        test_ntlm_button = tb.Button(ntlm_buttons_frame, text="Probar Credenciales", command=self.test_ntlm_credentials, bootstyle="info")
        test_ntlm_button.pack(side=tk.LEFT, padx=(0, 5))
        
        # Clear NTLM Credentials Button
        clear_ntlm_button = tb.Button(ntlm_buttons_frame, text="Limpiar Credenciales", command=self.clear_ntlm_credentials, bootstyle="warning")
        clear_ntlm_button.pack(side=tk.LEFT)

        # Automated Process Scheduling
        schedule_frame = tb.LabelFrame(settings_frame, text="Programación de Proceso Automatizado", padding="10")
        schedule_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        schedule_frame.grid_columnconfigure(1, weight=1)

        self.schedule_enabled_var = tk.BooleanVar()
        schedule_cb = tb.Checkbutton(schedule_frame, 
                                     text="Activar programación diaria",
                                     variable=self.schedule_enabled_var,
                                     command=self._toggle_schedule_inputs,
                                     bootstyle="primary")
        schedule_cb.grid(row=0, column=0, columnspan=3, sticky=tk.W, pady=(0, 10))

        tb.Label(schedule_frame, text="Hora de ejecución (formato 24h):").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        
        time_frame = tb.Frame(schedule_frame)
        time_frame.grid(row=1, column=1, sticky=tk.W)

        self.schedule_hour_var = tk.StringVar(value="02")
        self.schedule_minute_var = tk.StringVar(value="00")

        hour_entry = tb.Entry(time_frame, textvariable=self.schedule_hour_var, width=4)
        hour_entry.pack(side=tk.LEFT)
        tb.Label(time_frame, text=":").pack(side=tk.LEFT, padx=2)
        minute_entry = tb.Entry(time_frame, textvariable=self.schedule_minute_var, width=4)
        minute_entry.pack(side=tk.LEFT)
        
        save_schedule_button = tb.Button(schedule_frame, text="Guardar Programación", command=self.save_schedule, bootstyle="info")
        save_schedule_button.grid(row=1, column=2, padx=10)

        self.schedule_status_label = tb.Label(schedule_frame, text="La programación está desactivada.", bootstyle="secondary")
        self.schedule_status_label.grid(row=2, column=0, columnspan=3, sticky=tk.W, pady=(10, 0))

        # Configure tab expansion
        self.settings_tab.grid_columnconfigure(0, weight=1)
        self.settings_tab.grid_rowconfigure(2, weight=1)

    def _toggle_schedule_inputs(self):
        """Enable or disable time inputs based on the checkbox."""
        state = tk.NORMAL if self.schedule_enabled_var.get() else tk.DISABLED
        # Assuming hour_entry and minute_entry are children of time_frame
        for widget in self.settings_tab.winfo_children():
            if isinstance(widget, tb.LabelFrame) and "Programación" in widget.cget("text"):
                for time_frame in widget.winfo_children():
                    if isinstance(time_frame, tb.Frame):
                        for entry in time_frame.winfo_children():
                            if isinstance(entry, tb.Entry):
                                entry.configure(state=state)

    def select_download_dir(self):
        directory = filedialog.askdirectory(initialdir=self.dir_var.get())
        if directory:
            self.dir_var.set(directory)

    def update_speed_settings(self):
        """Update advanced settings based on selected speed preset"""
        preset = self.speed_preset.get()
        
        # Configure settings based on preset
        if preset == "slow":
            self.workers_var.set("1")
            self.delay_var.set("2")
            self.chunk_var.set("4096")
        elif preset == "medium":
            self.workers_var.set("3")
            self.delay_var.set("1")
            self.chunk_var.set("8192")
        elif preset == "fast":
            self.workers_var.set("5")
            self.delay_var.set("0.5")
            self.chunk_var.set("16384")
        elif preset == "ultra":
            self.workers_var.set("8")
            self.delay_var.set("0.1")
            self.chunk_var.set("32768")

    def _format_ntlm_status(self, user: Optional[str]) -> str:
        if user:
            return f"Cuenta NTLM guardada: {user}"
        return "Sin credenciales NTLM guardadas"

    def _update_ntlm_status_label(self) -> None:
        saved_user, _ = credentials_manager.get_credentials()
        status = self._format_ntlm_status(saved_user)
        if hasattr(self, "ntlm_status_label"):
            self.ntlm_status_label.configure(text=status)

    def save_ntlm_credentials(self):
        """Guarda las credenciales NTLM en el archivo local."""
        user = self.ntlm_user_var.get().strip()
        password = self.ntlm_pass_var.get()
        
        if not user or not password:
            messagebox.showwarning("Advertencia", "El usuario y la contraseña no pueden estar vacíos.")
            return
        
        self.ntlm_user_var.set(user)
        
        if credentials_manager.save_credentials(user, password):
            logger.info(f"Credenciales NTLM guardadas localmente para el usuario: {user}")
            self._update_ntlm_status_label()
            messagebox.showinfo(
                "Éxito",
                "Las credenciales NTLM se han guardado correctamente en el archivo local.\n"
                "Estas credenciales estarán disponibles cada vez que ejecute la aplicación."
            )
        else:
            logger.error("Error al guardar credenciales NTLM")
            messagebox.showerror(
                "Error",
                "No se pudieron guardar las credenciales. Revise los logs para más detalles."
            )

    def test_ntlm_credentials(self):
        """Realiza una solicitud de prueba contra el ReportServer para validar las credenciales."""
        user = self.ntlm_user_var.get().strip()
        password = self.ntlm_pass_var.get()
        
        if not user or not password:
            messagebox.showwarning("Advertencia", "Ingrese usuario y contraseña antes de probar.")
            return
        
        self.ntlm_user_var.set(user)
        test_url = os.getenv('REPORT_BASE_URL', BASE_URL)
        auth = HttpNtlmAuth(user, password)
        
        try:
            logger.info(f"Probando credenciales NTLM con usuario: {user} en {test_url}")
            response = requests.get(
                test_url,
                auth=auth,
                timeout=30,
                allow_redirects=False
            )
            status = response.status_code
            response.close()
            
            if status == 401:
                messagebox.showerror(
                    "Credenciales rechazadas",
                    "El servidor devolvió 401 (Unauthorized).\n"
                    "Verifica el dominio (ej. DOMINIO\\usuario) y que la contraseña sea correcta."
                )
            else:
                messagebox.showinfo(
                    "Autenticación exitosa",
                    f"El servidor respondió con HTTP {status}. Las credenciales parecen válidas."
                )
        except requests.exceptions.RequestException as exc:
            logger.error(f"Error probando credenciales NTLM: {exc}")
            messagebox.showerror(
                "Error de prueba",
                f"No se pudo validar las credenciales: {exc}"
            )

    def clear_ntlm_credentials(self):
        """Elimina las credenciales NTLM guardadas."""
        if messagebox.askyesno(
            "Confirmar",
            "¿Está seguro de que desea eliminar las credenciales guardadas?\n"
            "Esto requerirá que las configure nuevamente la próxima vez que use la aplicación."
        ):
            if credentials_manager.clear_credentials():
                # Limpiar los campos en la interfaz
                self.ntlm_user_var.set('')
                self.ntlm_pass_var.set('')
                self._update_ntlm_status_label()
                
                logger.info("Credenciales NTLM eliminadas")
                messagebox.showinfo("Éxito", "Las credenciales han sido eliminadas correctamente.")
            else:
                logger.error("Error al eliminar credenciales NTLM")
                messagebox.showerror("Error", "No se pudieron eliminar las credenciales. Revise los logs para más detalles.")

    def save_settings(self):
        try:
            # Validate inputs
            workers = int(self.workers_var.get())
            delay = float(self.delay_var.get())
            chunk_size = int(self.chunk_var.get())

            if workers < 1 or delay < 0 or chunk_size < 1024:
                raise ValueError("Valores inválidos en la configuración")
            
            # Update environment variables immediately
            os.environ['DOWNLOAD_MAX_WORKERS'] = str(workers)
            os.environ['DOWNLOAD_DELAY'] = str(delay)
            os.environ['CHUNK_SIZE'] = str(chunk_size)
            os.environ['DESCARGAS_DIR'] = self.dir_var.get()

            # Save to environment variables
            os.environ['DOWNLOAD_MAX_WORKERS'] = str(workers)
            os.environ['DOWNLOAD_DELAY'] = str(delay)
            os.environ['CHUNK_SIZE'] = str(chunk_size)
            os.environ['DESCARGAS_DIR'] = self.dir_var.get()

            # Create download directory if it doesn't exist
            os.makedirs(self.dir_var.get(), exist_ok=True)

            messagebox.showinfo("Éxito", "Configuración guardada correctamente")
        except ValueError as e:
            messagebox.showerror("Error", str(e))
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar la configuración: {str(e)}")

    def on_closing(self):
        """Handle window closing"""
        if self.is_running:
            if messagebox.askokcancel("Salir", "Hay un proceso en ejecución. ¿Deseas cancelarlo y salir?"):
                self.cancel_process()
                self.destroy()
        else:
            self.destroy()

    def setup_automated_tab(self):
        """Configura la pestaña de proceso automatizado que combina descarga, cruce, formateo y optimización"""
        # Título principal
        title_frame = tb.Frame(self.automated_tab)
        title_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), padx=10, pady=10)
        
        title_label = tb.Label(title_frame, text="🚀 PROCESO AUTOMATIZADO COMPLETO", 
                               font=('Helvetica', 16, 'bold'))
        title_label.pack()
        
        subtitle_label = tb.Label(title_frame, 
                                  text="Descarga → Procesar → Cruce → Formateo → Conversion", 
                                  font=('Helvetica', 10, 'italic'))
        subtitle_label.pack(pady=(5, 0))

        # Frame de configuración de paso inicial
        start_step_frame = tb.LabelFrame(self.automated_tab, text="Configuración de Inicio", padding="15")
        start_step_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), padx=10, pady=5)
        start_step_frame.grid_columnconfigure(0, weight=1)

        step_info_label = tb.Label(start_step_frame, text="Selecciona desde qué paso quieres iniciar:", 
                                   font=('Helvetica', 10, 'bold'))
        step_info_label.pack(anchor=tk.W, pady=(0, 10))

        # Variable para el paso inicial
        self.start_step_var = tk.StringVar(value="step1")
        
        # Radio buttons para seleccionar paso inicial
        step_options_frame = tb.Frame(start_step_frame)
        step_options_frame.pack(fill=tk.X, pady=5)
        
        step1_radio = tb.Radiobutton(step_options_frame, 
                                     text="📥 Paso 1: Comenzar con DESCARGA (proceso completo desde cero)",
                                     variable=self.start_step_var, 
                                     value="step1",
                                     bootstyle="primary")
        step1_radio.pack(anchor=tk.W, pady=2)
        
        step2_radio = tb.Radiobutton(step_options_frame, 
                                     text="⚙️ Paso 2: Comenzar con PROCESAMIENTO (saltar descarga, usar archivos existentes)",
                                     variable=self.start_step_var, 
                                     value="step2",
                                     bootstyle="primary")
        step2_radio.pack(anchor=tk.W, pady=2)

        # Frame de información
        info_frame = tb.LabelFrame(self.automated_tab, text="Información del Proceso", padding="15")
        info_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), padx=10, pady=5)
        info_frame.grid_columnconfigure(0, weight=1)

        info_text = """Este proceso automatizado ejecutará secuencialmente:

1. 📥 DESCARGA: Descarga archivos CCM y PRR según configuración
2. ⚙️ PROCESAR ARCHIVOS: Extrae y procesa datos de archivos descargados
3. 🔄 CRUCE: Procesamiento optimizado de cruces de datos
4. 📊 FORMATEO: Formateo avanzado con ultra-threading
5. ⚡ CONVERSION: Optimización final para visualización

⚠️ IMPORTANTE: 
• Si seleccionas "Paso 1", se ejecutará el proceso completo desde cero
• Si seleccionas "Paso 2", asegúrate de que existan archivos consolidados previos
• Verifica que la configuración de módulos esté correcta antes de iniciar"""

        info_label = tb.Label(info_frame, text=info_text, justify=tk.LEFT, wraplength=800)
        info_label.pack(padx=10, pady=5)

        # Progress section
        auto_progress_frame = tb.LabelFrame(self.automated_tab, text="Estado del Proceso Automatizado", padding="10")
        auto_progress_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), padx=10, pady=5)
        auto_progress_frame.grid_columnconfigure(0, weight=1)
        
        self.auto_progress_label = tb.Label(auto_progress_frame, text="Esperando inicio...")
        self.auto_progress_label.grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        
        self.auto_progress_bar = tb.Progressbar(auto_progress_frame, mode='determinate', bootstyle="success-striped")
        self.auto_progress_bar.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 5))

        # Log area for automated process
        auto_log_frame = tb.LabelFrame(self.automated_tab, text="Registro del Proceso Automatizado", padding="10")
        auto_log_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), padx=10, pady=5)
        auto_log_frame.grid_columnconfigure(0, weight=1)
        auto_log_frame.grid_rowconfigure(0, weight=1)

        self.auto_log_area = scrolledtext.ScrolledText(auto_log_frame, height=15, wrap=tk.WORD)
        self.auto_log_area.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5, pady=5)

        # Control button
        auto_button_frame = tb.Frame(self.automated_tab)
        auto_button_frame.grid(row=5, column=0, columnspan=2, pady=20)

        self.auto_start_button = tb.Button(auto_button_frame, text="🚀 INICIAR PROCESO AUTOMATIZADO COMPLETO",
                                           command=self.start_automated_process,
                                           bootstyle="success",
                                           width=40)
        self.auto_start_button.pack(pady=10)

        # Configure grid weights for proper expansion
        self.automated_tab.grid_columnconfigure(0, weight=1)
        self.automated_tab.grid_rowconfigure(4, weight=1)

    def start_automated_process(self):
        """Inicia el proceso automatizado completo"""
        if not self.is_running:
            # Validar prerrequisitos según el paso inicial
            start_step = self.start_step_var.get()
            if start_step == "step1":
                # Proceso completo desde descarga
                if not self._validate_process_prerequisites():
                    return
            else:
                # Proceso desde procesamiento
                if not self._validate_processing_prerequisites():
                    return

            self.is_running = True
            self._disable_all_buttons()
            self.auto_start_button.configure(state=tk.DISABLED)
            self.auto_progress_bar["value"] = 0
            self.auto_log_area.delete(1.0, tk.END)
            
            # Determinar paso inicial
            start_step = self.start_step_var.get()
            if start_step == "step1":
                self.queue.put(("auto_log", "🚀 INICIANDO PROCESO AUTOMATIZADO COMPLETO (desde Paso 1: Descarga)"))
                total_steps = 5
                step_start = 1
            else:
                self.queue.put(("auto_log", "🚀 INICIANDO PROCESO AUTOMATIZADO (desde Paso 2: Procesamiento)"))
                self.queue.put(("auto_log", "⚠️ Saltando descarga - usando archivos consolidados existentes"))
                total_steps = 4
                step_start = 2

            def run_process():
                try:
                    current_step = step_start
                    
                    # Paso 1: Descarga (solo si se seleccionó comenzar desde paso 1)
                    if start_step == "step1":
                        self.queue.put(("auto_log", "📥 PASO 1/5: Iniciando descarga de archivos CCM y PRR..."))
                        self.queue.put(("auto_progress_label", "Paso 1/5: Descargando archivos..."))
                        self.queue.put(("auto_progress", 4))
                        
                        # Mostrar configuración de workers
                        workers_count = os.getenv('DOWNLOAD_MAX_WORKERS', '2')
                        delay_config = os.getenv('DOWNLOAD_DELAY', '0.5')
                        direct_download = os.getenv('DIRECT_DOWNLOAD', 'false')
                        self.queue.put(("auto_log", f"   ⚙️ Configuración: {workers_count} workers, {delay_config}s delay, descarga directa: {direct_download}"))
                        
                        # Usar la función de descarga con callbacks
                        def auto_log_callback(message):
                            self.queue.put(("auto_log", f"   📥 {message}"))
                        
                        def auto_error_callback(file_path, error_message):
                            self.queue.put(("auto_log", f"   ❌ Error en {file_path}: {error_message}"))
                        
                        def auto_progress_callback(progress):
                            # Mapear progreso de descarga (0-100) a rango 4-20
                            mapped_progress = 4 + (progress * 0.16)
                            self.queue.put(("auto_progress", mapped_progress))
                        
                        # Actualizar variables de entorno con configuración actual (igual que proceso individual)
                        os.environ['DOWNLOAD_MAX_WORKERS'] = self.workers_var.get()
                        os.environ['DOWNLOAD_DELAY'] = self.delay_var.get()
                        os.environ['CHUNK_SIZE'] = self.chunk_var.get()
                        os.environ['DIRECT_DOWNLOAD'] = str(self.direct_download_var.get()).lower()
                        
                        # Ejecutar descarga
                        selected_modules = {name: var.get() for name, var in self.module_vars.items()}
                        download_result = descargar_y_consolidar(
                            download_option=self.download_option.get(),
                            progress_callback=auto_progress_callback,
                            log_callback=auto_log_callback,
                            error_callback=auto_error_callback,
                            selected_modules=selected_modules,
                            overwrite=True  # CRÍTICO: Igual que proceso individual
                        )
                        
                        if download_result is None or download_result is False:
                            self.queue.put(("auto_error", "Error en la descarga de archivos - proceso falló o no retornó valor"))
                            return
                        
                        self.queue.put(("auto_log", "✅ PASO 1 COMPLETADO: Descarga exitosa"))
                        self.queue.put(("auto_progress", 20))
                        current_step = 2
                    
                    # Calcular progreso base según si saltamos descarga
                    if start_step == "step1":
                        # Proceso completo 5 pasos: 20%, 40%, 60%, 80%, 100%
                        progress_steps = [20, 40, 60, 80, 100]
                        step_labels = ["2/5", "3/5", "4/5", "5/5"]
                    else:
                        # Proceso desde paso 2 (4 pasos): 25%, 50%, 75%, 100%
                        progress_steps = [25, 50, 75, 100]
                        step_labels = ["1/4", "2/4", "3/4", "4/4"]
                        self.queue.put(("auto_progress", 0))
                    
                    # Índice para arrays de progreso
                    progress_idx = 0 if start_step == "step2" else 1
                    
                    # Paso 2: Procesar Archivos
                    self.queue.put(("auto_log", f"⚙️ PASO {current_step}/{total_steps}: Iniciando procesamiento de archivos..."))
                    self.queue.put(("auto_progress_label", f"Paso {current_step}/{total_steps}: Procesando archivos descargados..."))
                    
                    processor = ExcelProcessor()
                    successful_files = processor.process_all_files(
                        lambda msg: self.queue.put(("auto_log", f"   ⚙️ {msg}"))
                    )
                    
                    if not successful_files:
                        self.queue.put(("auto_error", "Error en el procesamiento de archivos"))
                        return
                    
                    self.queue.put(("auto_log", f"✅ PASO {current_step} COMPLETADO: Procesamiento de archivos exitoso"))
                    self.queue.put(("auto_progress", progress_steps[progress_idx]))
                    current_step += 1
                    progress_idx += 1
                    
                    # Paso 3: ULTRA Cruce
                    self.queue.put(("auto_log", f"🔄 PASO {current_step}/{total_steps}: Iniciando ULTRA Cruce..."))
                    self.queue.put(("auto_progress_label", f"Paso {current_step}/{total_steps}: Procesando ULTRA Cruce..."))
                    
                    successful_files = processor.ultra_fast_cross_processing(
                        lambda msg: self.queue.put(("auto_log", f"   🔄 {msg}"))
                    )
                    
                    if not successful_files:
                        self.queue.put(("auto_error", "Error en el procesamiento de cruces"))
                        return
                    
                    self.queue.put(("auto_log", f"✅ PASO {current_step} COMPLETADO: ULTRA Cruce exitoso"))
                    self.queue.put(("auto_progress", progress_steps[progress_idx]))
                    current_step += 1
                    progress_idx += 1
                    
                    # Paso 4: ESTABLE Formateo
                    self.queue.put(("auto_log", f"📊 PASO {current_step}/{total_steps}: Iniciando ESTABLE Formateo..."))
                    self.queue.put(("auto_progress_label", f"Paso {current_step}/{total_steps}: Ejecutando ESTABLE Formateo..."))
                    
                    successful_files = processor.ultra_threaded_format_files(
                        lambda msg: self.queue.put(("auto_log", f"   📊 {msg}"))
                    )
                    
                    if not successful_files:
                        self.queue.put(("auto_error", "Error en el formateo de archivos"))
                        return
                    
                    self.queue.put(("auto_log", f"✅ PASO {current_step} COMPLETADO: ESTABLE Formateo exitoso"))
                    self.queue.put(("auto_progress", progress_steps[progress_idx]))
                    current_step += 1
                    progress_idx += 1
                    
                    # Paso 5: Optimización para Streamlit
                    self.queue.put(("auto_log", f"⚡ PASO {current_step}/{total_steps}: Iniciando optimización para Streamlit..."))
                    self.queue.put(("auto_progress_label", f"Paso {current_step}/{total_steps}: Optimizando para Streamlit..."))
                    
                    # Ejecutar optimización directamente sin diálogo
                    try:
                        from optimize_files import analyze_file_structure, optimize_dataframe, save_in_multiple_formats
                        import glob
                        from pathlib import Path
                        
                        # Buscar archivos consolidados
                        ccm_file = "descargas/CCM/consolidado_final_CCM_personal.xlsx"
                        prr_file = "descargas/PRR/consolidado_final_PRR_personal.xlsx"
                        
                        files_to_process = []
                        if os.path.exists(ccm_file):
                            files_to_process.append((ccm_file, "CCM"))
                        if os.path.exists(prr_file):
                            files_to_process.append((prr_file, "PRR"))
                        
                        if files_to_process:
                            output_dir = "optimized"
                            Path(output_dir).mkdir(exist_ok=True)
                            
                            for file_path, file_type in files_to_process:
                                self.queue.put(("auto_log", f"   📊 Analizando archivo {file_type}..."))
                                
                                # Analizar estructura
                                df, analysis_info = analyze_file_structure(file_path)
                                if df is None:
                                    self.queue.put(("auto_log", f"   ❌ Error al leer archivo {file_type}"))
                                    continue
                                
                                self.queue.put(("auto_log", f"   📏 {df.shape[0]:,} filas × {df.shape[1]} columnas"))
                                self.queue.put(("auto_log", f"   💾 Tamaño: {analysis_info['memory_usage_mb']:.1f} MB"))
                                
                                # Optimizar DataFrame
                                self.queue.put(("auto_log", f"   🔧 Optimizando {file_type}..."))
                                df_optimized = optimize_dataframe(df, analysis_info)
                                
                                # Guardar solo en formato Pickle.GZ (más rápido)
                                self.queue.put(("auto_log", f"   💾 Guardando {file_type} optimizado..."))
                                formats_info = save_in_multiple_formats(
                                    df_optimized, 
                                    f"consolidado_final_{file_type}_personal", 
                                    output_dir, 
                                    produce_only_pickle_gz=True
                                )
                                
                                # Mostrar resultados
                                if 'pickle_gz' in formats_info:
                                    original_size = analysis_info['original_size_mb']
                                    optimized_size = formats_info['pickle_gz']['size_mb']
                                    reduction = (1 - optimized_size / original_size) * 100
                                    self.queue.put(("auto_log", f"   ✅ {file_type}: {optimized_size:.2f} MB (-{reduction:.1f}%)"))
                            
                            self.queue.put(("auto_log", f"✅ PASO {current_step} COMPLETADO: Optimización para Streamlit exitosa"))
                            self.queue.put(("auto_log", f"   📁 Archivos optimizados en: {Path(output_dir).resolve()}"))
                        else:
                            self.queue.put(("auto_log", "⚠️ No se encontraron archivos consolidados para optimizar"))
                    
                    except ImportError:
                        self.queue.put(("auto_log", "⚠️ Módulo de optimización no disponible, saltando paso final"))
                    except Exception as e:
                        self.queue.put(("auto_log", f"⚠️ Error en optimización: {str(e)}"))
                    
                    self.queue.put(("auto_progress", 100))
                    
                    # Proceso completado
                    self.queue.put(("auto_log", ""))
                    self.queue.put(("auto_log", "🎉 ¡PROCESO AUTOMATIZADO COMPLETADO EXITOSAMENTE!"))
                    if start_step == "step1":
                        self.queue.put(("auto_log", "✅ Todos los 5 pasos han sido ejecutados correctamente"))
                    else:
                        self.queue.put(("auto_log", "✅ Todos los 4 pasos han sido ejecutados correctamente"))
                    self.queue.put(("auto_progress_label", "¡Proceso completado exitosamente!"))
                    
                except Exception as e:
                    self.queue.put(("auto_error", f"Error fatal en proceso automatizado: {str(e)}"))
                finally:
                    self.queue.put(("auto_finished", None))

            self.process_thread = threading.Thread(target=run_process)
            self.process_thread.start()
            self.check_auto_queue()

    def _disable_all_buttons(self):
        """Deshabilita todos los botones de procesamiento"""
        self.process_button.configure(state=tk.DISABLED)
        self.ultra_cross_button.configure(state=tk.DISABLED)
        self.max_perf_format_button.configure(state=tk.DISABLED)
        self.optimize_button.configure(state=tk.DISABLED)
        
        # Deshabilitar botones de descarga también
        self.start_button.configure(state=tk.DISABLED)
        self.pause_button.configure(state=tk.DISABLED)
        self.cancel_button.configure(state=tk.DISABLED)
        
        # Deshabilitar botón de proceso automatizado
        self.auto_start_button.configure(state=tk.DISABLED)

    def _enable_all_buttons(self):
        """Habilita todos los botones de procesamiento"""
        self.process_button.configure(state=tk.NORMAL)
        self.ultra_cross_button.configure(state=tk.NORMAL)
        self.max_perf_format_button.configure(state=tk.NORMAL)
        self.optimize_button.configure(state=tk.NORMAL)
        
        # Habilitar botones de descarga también
        self.start_button.configure(state=tk.NORMAL)
        self.pause_button.configure(state=tk.DISABLED)  # Mantener pausar deshabilitado por defecto
        self.cancel_button.configure(state=tk.DISABLED)  # Mantener cancelar deshabilitado por defecto
        
        # Habilitar botón de proceso automatizado
        self.auto_start_button.configure(state=tk.NORMAL)

    def start_optimized_cross_processing(self):
        """Inicia el procesamiento de cruces optimizado"""
        if not self.is_running:
            self.is_running = True
            self._disable_all_buttons()
            self.process_progress_bar["value"] = 0
            self.process_log_area.delete(1.0, tk.END)
            self.log_message("🚀 Iniciando procesamiento de cruces OPTIMIZADO...", process=True)

            def run_process():
                try:
                    processor = ExcelProcessor()
                    successful_files = processor.optimized_cross_processing(
                        lambda msg: self.queue.put(("process_log", msg))
                    )
                    if not successful_files:
                        self.queue.put(("process_error", "No se procesó ningún cruce correctamente"))
                    else:
                        self.queue.put(("process_log", f"🎉 Procesamiento optimizado completado exitosamente"))
                except Exception as e:
                    self.queue.put(("process_error", str(e)))
                finally:
                    self.queue.put(("process_finished", None))

            self.process_thread = threading.Thread(target=run_process)
            self.process_thread.start()
            self.check_queue()

    def start_optimized_format_processing(self):
        """Inicia el formateo optimizado"""
        if not self.is_running:
            self.is_running = True
            self._disable_all_buttons()
            self.process_progress_bar["value"] = 0
            self.process_log_area.delete(1.0, tk.END)
            self.log_message("⚡ Iniciando formateo OPTIMIZADO...", process=True)

            def run_process():
                try:
                    processor = ExcelProcessor()
                    successful_files = processor.optimized_format_files(
                        lambda msg: self.queue.put(("process_log", msg))
                    )
                    if not successful_files:
                        self.queue.put(("process_error", "No se pudo formatear ningún archivo"))
                    else:
                        self.queue.put(("process_log", f"⚡ Formateo optimizado completado exitosamente"))
                        # Auto-actualizar barra de progreso
                        for i in range(0, 101, 20):
                            self.queue.put(("process_progress", i))
                except Exception as e:
                    self.queue.put(("process_error", str(e)))
                finally:
                    self.queue.put(("process_finished", None))

            self.process_thread = threading.Thread(target=run_process)
            self.process_thread.start()
            self.check_queue()

    def start_ultra_format_processing(self):
        """Inicia el formateo ULTRA"""
        if not self.is_running:
            self.is_running = True
            self._disable_all_buttons()
            self.process_progress_bar["value"] = 0
            self.process_log_area.delete(1.0, tk.END)
            self.log_message("🚀 Iniciando formateo ULTRA...", process=True)

            def run_process():
                try:
                    processor = ExcelProcessor()
                    successful_files = processor.ultra_optimized_format_files(
                        lambda msg: self.queue.put(("process_log", msg))
                    )
                    if not successful_files:
                        self.queue.put(("process_error", "No se pudo formatear ningún archivo"))
                    else:
                        self.queue.put(("process_log", f"🎉 Formateo ULTRA completado exitosamente"))
                        # Auto-actualizar barra de progreso
                        for i in range(0, 101, 25):
                            self.queue.put(("process_progress", i))
                except Exception as e:
                    self.queue.put(("process_error", str(e)))
                finally:
                    self.queue.put(("process_finished", None))

            self.process_thread = threading.Thread(target=run_process)
            self.process_thread.start()
            self.check_queue()

    def start_ultra_fast_format_processing(self):
        """Inicia el formateo ULTRA-FAST"""
        if not self.is_running:
            self.is_running = True
            self._disable_all_buttons()
            self.process_progress_bar["value"] = 0
            self.process_log_area.delete(1.0, tk.END)
            self.log_message("🚀 Iniciando formateo ULTRA-FAST...", process=True)

            def run_process():
                try:
                    processor = ExcelProcessor()
                    successful_files = processor.ultra_fast_format_files(
                        lambda msg: self.queue.put(("process_log", msg))
                    )
                    if not successful_files:
                        self.queue.put(("process_error", "No se pudo formatear ningún archivo"))
                    else:
                        self.queue.put(("process_log", f"🎉 Formateo ULTRA-FAST completado exitosamente"))
                        # Auto-actualizar barra de progreso
                        for i in range(0, 101, 25):
                            self.queue.put(("process_progress", i))
                except Exception as e:
                    self.queue.put(("process_error", str(e)))
                finally:
                    self.queue.put(("process_finished", None))

            self.process_thread = threading.Thread(target=run_process)
            self.process_thread.start()
            self.check_queue()

    def start_max_performance_format_processing(self):
        """Inicia el formateo de MÁXIMO RENDIMIENTO con multiprocessing"""
        if not self.is_running:
            self.is_running = True
            self._disable_all_buttons()
            self.process_progress_bar["value"] = 0
            self.process_log_area.delete(1.0, tk.END)
            self.log_message("🔥 Iniciando formateo ESTABLE (ultra-threading)...", process=True)

            def run_process():
                try:
                    processor = ExcelProcessor()
                    successful_files = processor.ultra_threaded_format_files(
                        lambda msg: self.queue.put(("process_log", msg))
                    )
                    if not successful_files:
                        self.queue.put(("process_error", "No se pudo formatear ningún archivo"))
                    else:
                        self.queue.put(("process_log", f"🔥 Formateo de MÁXIMO RENDIMIENTO completado exitosamente"))
                        # Auto-actualizar barra de progreso
                        for i in range(0, 101, 33):
                            self.queue.put(("process_progress", i))
                except Exception as e:
                    self.queue.put(("process_error", str(e)))
                finally:
                    self.queue.put(("process_finished", None))

            self.process_thread = threading.Thread(target=run_process)
            self.process_thread.start()
            self.check_queue()

    def start_regenerate_personal(self):
        """Inicia la regeneración de archivos personal"""
        if not self.is_running:
            self.is_running = True
            self._disable_all_buttons()
            self.process_progress_bar["value"] = 0
            self.process_log_area.delete(1.0, tk.END)
            self.log_message("🔄 Iniciando regeneración de archivos personal...", process=True)

            def run_process():
                try:
                    processor = ExcelProcessor()
                    successful_files = processor.regenerate_personal_files(
                        lambda msg: self.queue.put(("process_log", msg))
                    )
                    if not successful_files:
                        self.queue.put(("process_error", "No se pudo regenerar ningún archivo personal"))
                    else:
                        self.queue.put(("process_log", f"🎉 Regeneración de archivos personal completada exitosamente"))
                except Exception as e:
                    self.queue.put(("process_error", str(e)))
                finally:
                    self.queue.put(("process_finished", None))

            self.process_thread = threading.Thread(target=run_process)
            self.process_thread.start()
            self.check_queue()

    def start_ultra_fast_cross_processing(self):
        """Inicia el procesamiento de cruces ULTRA-OPTIMIZADO"""
        if not self.is_running:
            # La validación crítica de PERSONAL.xlsx ahora se hace al inicio.
            # Ya no se necesita una validación específica aquí.
            self.is_running = True
            # ... (el resto del método se mantiene igual)
            
    def check_auto_queue(self):
        """Verificar cola específica para proceso automatizado"""
        try:
            while True:
                item = self.queue.get_nowait()
                message_type = item[0]
                message_data = item[1]
                
                if message_type == "auto_log":
                    self.auto_log_message(message_data)
                elif message_type == "auto_error":
                    self.auto_log_message(message_data, error=True)
                elif message_type == "auto_progress":
                    self.auto_progress_bar["value"] = message_data
                elif message_type == "auto_progress_label":
                    self.auto_progress_label.configure(text=message_data)
                elif message_type == "auto_finished":
                    self.auto_process_finished()
                    break
        except queue.Empty:
            pass
        
        if self.is_running:
            self.after(100, self.check_auto_queue)

    def auto_log_message(self, message, error=False):
        """Añadir mensaje al área de log de proceso automatizado"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        formatted_message = f"[{timestamp}] {message}\n"
        
        self.auto_log_area.insert(tk.END, formatted_message)
        if error:
            # Aplicar color rojo para errores
            start_line = self.auto_log_area.index(tk.END + " -1c linestart")
            end_line = self.auto_log_area.index(tk.END + " -1c")
            self.auto_log_area.tag_add("error", start_line, end_line)
            self.auto_log_area.tag_configure("error", foreground="red")
        
        self.auto_log_area.see(tk.END)

    def auto_process_finished(self):
        """Finalizar proceso automatizado"""
        self.is_running = False
        self._enable_all_buttons()
        self.auto_start_button.configure(state=tk.NORMAL)
        self.auto_log_message("🏁 Proceso automatizado finalizado")

    def _validate_process_prerequisites(self) -> bool:
        """
        Valida que existan todos los prerrequisitos para el proceso de descarga
        """
        try:
            # Validar carpetas de descarga
            required_dirs = ['descargas', 'descargas/CCM', 'descargas/PRR']
            missing_dirs = []
            
            for dir_path in required_dirs:
                if not os.path.exists(dir_path):
                    missing_dirs.append(dir_path)
            
            if missing_dirs:
                # Intentar crear las carpetas faltantes
                try:
                    for dir_path in missing_dirs:
                        os.makedirs(dir_path, exist_ok=True)
                    self.log_message(f"✅ Carpetas creadas automáticamente: {', '.join(missing_dirs)}")
                except Exception as e:
                    messagebox.showerror(
                        "Error de Estructura", 
                        f"No se pudieron crear las carpetas necesarias:\n{', '.join(missing_dirs)}\n\nError: {str(e)}"
                    )
                    return False
            
            # Validar configuración de entorno
            required_env_vars = ['DOWNLOAD_MAX_WORKERS', 'DOWNLOAD_DELAY', 'CHUNK_SIZE']
            missing_env = []
            
            for var in required_env_vars:
                if not os.getenv(var):
                    missing_env.append(var)
            
            if missing_env:
                # Configurar valores por defecto
                defaults = {
                    'DOWNLOAD_MAX_WORKERS': '3',
                    'DOWNLOAD_DELAY': '1.0', 
                    'CHUNK_SIZE': '8192'
                }
                for var in missing_env:
                    os.environ[var] = defaults.get(var, '1')
                self.log_message(f"⚙️ Variables de entorno configuradas automáticamente")
                
            return True
            
        except Exception as e:
            logger.error(f"Error validando prerrequisitos de descarga: {str(e)}")
            messagebox.showerror(
                "Error de Validación",
                f"No se pudieron validar los prerrequisitos para la descarga:\n{str(e)}"
            )
            return False

    def _validate_processing_prerequisites(self) -> bool:
        """
        Valida que existan los archivos necesarios para el procesamiento
        """
        try:
            # Buscar archivos consolidados para procesar
            modules = ['CCM', 'PRR']
            available_files = []
            missing_files = []
            
            for module in modules:
                # Buscar tanto archivos CSV consolidados como Excel consolidados
                csv_file = f"descargas/{module}/consolidado_total_{module}.csv"
                xlsx_file = f"descargas/{module}/consolidado_final_{module}_personal.xlsx"
                
                if os.path.exists(csv_file) or os.path.exists(xlsx_file):
                    available_files.append(module)
                else:
                    missing_files.append(module)
            
            if not available_files:
                messagebox.showwarning(
                    "Archivos No Encontrados",
                    "No se encontraron archivos consolidados para procesar.\n\n"
                    "Debes ejecutar primero el proceso de descarga o verificar que "
                    "existan archivos consolidados en las carpetas descargas/CCM/ y descargas/PRR/.\n\n"
                    f"Archivos buscados:\n"
                    f"• consolidado_total_[MODULE].csv\n"
                    f"• consolidado_final_[MODULE]_personal.xlsx"
                )
                return False
            
            if missing_files:
                response = messagebox.askyesno(
                    "Archivos Parciales Encontrados",
                    f"Se encontraron archivos para: {', '.join(available_files)}\n"
                    f"Faltantes: {', '.join(missing_files)}\n\n"
                    f"¿Deseas continuar procesando solo los archivos disponibles?"
                )
                if not response:
                    return False
            
            self.log_message(f"✅ Archivos disponibles para procesamiento: {', '.join(available_files)}", process=True)
            return True
            
        except Exception as e:
            logger.error(f"Error validando prerrequisitos de procesamiento: {str(e)}")
            messagebox.showerror(
                "Error de Validación",
                f"No se pudieron validar los prerrequisitos para el procesamiento:\n{str(e)}"
            )
            return False

    def _validate_cross_processing_prerequisites(self) -> bool:
        """
        Valida prerrequisitos específicos para procesos de cruce
        """
        try:
            # Verificar archivo PERSONAL.xlsx para cruces
            personal_file = "ASIGNACIONES/PERSONAL.xlsx"
            
            if not os.path.exists(personal_file):
                response = messagebox.askyesno(
                    "Archivo de Cruce Faltante",
                    f"El archivo '{personal_file}' es necesario para los procesos de cruce.\n\n"
                    f"¿Deseas continuar sin este archivo? (Los cruces pueden fallar)"
                )
                if not response:
                    return False
                else:
                    self.log_message("⚠️ Procesando sin archivo PERSONAL.xlsx - Los cruces pueden fallar", process=True)
            else:
                self.log_message("✅ Archivo PERSONAL.xlsx encontrado para cruces", process=True)
            
            return True
            
        except Exception as e:
            logger.error(f"Error validando prerrequisitos de cruce: {str(e)}")
            return True  # No es crítico, permitir continuar

    def save_schedule(self):
        """Valida y guarda la configuración de la programación."""
        if not self.schedule_enabled_var.get():
            config = {"enabled": False, "hour": -1, "minute": -1}
        else:
            try:
                hour = int(self.schedule_hour_var.get())
                minute = int(self.schedule_minute_var.get())
                if not (0 <= hour <= 23 and 0 <= minute <= 59):
                    raise ValueError("Hora o minuto fuera de rango.")
                
                config = {"enabled": True, "hour": hour, "minute": minute}

            except ValueError:
                messagebox.showerror("Error de Validación", "Por favor, introduce una hora (0-23) y minuto (0-59) válidos.")
                return

        with self.schedule_lock:
            self.scheduled_time = (config["hour"], config["minute"]) if config["enabled"] else None
            with open(self.schedule_file, 'w') as f:
                json.dump(config, f)
        
        self.update_schedule_status_label()
        messagebox.showinfo("Programación Guardada", "La configuración de programación ha sido guardada.")

    def load_schedule(self):
        """Carga la configuración de programación desde el archivo."""
        try:
            if os.path.exists(self.schedule_file):
                with open(self.schedule_file, 'r') as f:
                    config = json.load(f)
                
                with self.schedule_lock:
                    self.schedule_enabled_var.set(config.get("enabled", False))
                    self.schedule_hour_var.set(f"{config.get('hour', 2):02d}")
                    self.schedule_minute_var.set(f"{config.get('minute', 0):02d}")
                    if config.get("enabled"):
                        self.scheduled_time = (config.get("hour"), config.get("minute"))
            else:
                self.schedule_enabled_var.set(False)
        except (json.JSONDecodeError, KeyError):
            self.schedule_enabled_var.set(False)
        
        self.update_schedule_status_label()
        self._toggle_schedule_inputs()

    def update_schedule_status_label(self):
        """Actualiza la etiqueta de estado de la programación."""
        with self.schedule_lock:
            if self.schedule_enabled_var.get() and self.scheduled_time:
                hour, minute = self.scheduled_time
                self.schedule_status_label.config(text=f"Programado para ejecutarse diariamente a las {hour:02d}:{minute:02d}.", bootstyle="success")
            else:
                self.schedule_status_label.config(text="La programación está desactivada.", bootstyle="secondary")
                
    def start_scheduler_thread(self):
        """Inicia el hilo de fondo para el programador."""
        scheduler_thread = threading.Thread(target=self._scheduler_worker, daemon=True)
        scheduler_thread.start()

    def _scheduler_worker(self):
        """
        Hilo que se ejecuta en segundo plano para comprobar la hora
        y disparar el proceso automatizado.
        """
        logger.info("Scheduler thread started.")
        while True:
            time.sleep(60) # Revisar cada 60 segundos

            with self.schedule_lock:
                if not self.scheduled_time or self.is_running:
                    continue

                now = datetime.now()
                today = date.today()
                
                # Prevenir ejecuciones múltiples en el mismo día
                if self.last_run_date == today:
                    continue

                scheduled_hour, scheduled_minute = self.scheduled_time
                if now.hour == scheduled_hour and now.minute == scheduled_minute:
                    logger.info(f"¡Hora programada alcanzada! ({scheduled_hour:02d}:{scheduled_minute:02d}). Iniciando proceso automatizado.")
                    self.last_run_date = today
                    
                    # Usar self.after para asegurar que el proceso se inicie en el hilo principal de la GUI
                    self.after(0, self.start_automated_process)

if __name__ == "__main__":
    app = MainWindow()
    app.mainloop()
